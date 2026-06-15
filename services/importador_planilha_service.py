from openpyxl import load_workbook

from database.repositories import (
    obter_ou_criar_orgao,
    cadastrar_ou_atualizar_processo_planilha,
)


COLUNAS_OBRIGATORIAS = [
    "Empresa",
    "CNPJ",
    "Município",
    "Exercício",
    "Processo",
    "Código",
    "Acesso",
    "Login",
    "Senha",
]


def normalizar_valor(valor):
    if valor is None:
        return ""

    return str(valor).strip()


def identificar_chave_robo(acesso):
    acesso_normalizado = acesso.lower()

    if "consultaprotocolo.curitiba.pr.gov.br" in acesso_normalizado:
        return "curitiba"

    return None


def importar_planilha_base(caminho_planilha="Planilha_Base.xlsx"):
    workbook = load_workbook(
        caminho_planilha,
        data_only=True
    )

    planilha = workbook.active

    cabecalhos = [
        normalizar_valor(celula.value)
        for celula in planilha[1]
    ]

    for coluna in COLUNAS_OBRIGATORIAS:
        if coluna not in cabecalhos:
            raise ValueError(
                f"Coluna obrigatória não encontrada na planilha: {coluna}"
            )

    indices = {
        coluna: cabecalhos.index(coluna)
        for coluna in COLUNAS_OBRIGATORIAS
    }

    total_linhas = 0
    total_importados = 0
    total_ignorados = 0
    sem_robo = 0

    for linha in planilha.iter_rows(
        min_row=2,
        values_only=True
    ):

        total_linhas += 1

        empresa = normalizar_valor(linha[indices["Empresa"]])
        cnpj = normalizar_valor(linha[indices["CNPJ"]])
        municipio = normalizar_valor(linha[indices["Município"]])
        exercicio = normalizar_valor(linha[indices["Exercício"]])
        processo = normalizar_valor(linha[indices["Processo"]])
        codigo = normalizar_valor(linha[indices["Código"]])
        acesso = normalizar_valor(linha[indices["Acesso"]])
        login = normalizar_valor(linha[indices["Login"]])
        senha = normalizar_valor(linha[indices["Senha"]])

        if not processo or not acesso:
            total_ignorados += 1
            continue

        chave_robo = identificar_chave_robo(acesso)

        if chave_robo is None:
            sem_robo += 1

        nome_orgao = municipio or "Órgão não informado"

        orgao_id = obter_ou_criar_orgao(
            nome=nome_orgao,
            tipo="Prefeitura",
            url=acesso,
            chave_robo=chave_robo,
        )

        dados_processo = {
            "orgao_id": orgao_id,
            "empresa": empresa,
            "cnpj": cnpj,
            "municipio": municipio,
            "exercicio": exercicio,
            "numero_processo": processo,
            "codigo": codigo,
            "acesso": acesso,
            "login": login,
            "senha": senha,
        }

        cadastrar_ou_atualizar_processo_planilha(
            dados_processo
        )

        total_importados += 1

    return {
        "total_linhas": total_linhas,
        "total_importados": total_importados,
        "total_ignorados": total_ignorados,
        "sem_robo": sem_robo,
    }