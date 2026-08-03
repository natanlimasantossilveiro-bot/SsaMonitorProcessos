import json
import os
import subprocess
import sys
from datetime import date, datetime, timedelta
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from html import escape


def _fmt_data(val, hora=False):
    """Formata date/datetime para padrão brasileiro. hora=True inclui HH:MM."""
    if val is None or val == "" or val == "—":
        return "—"
    try:
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y - %H:%M") if hora else val.strftime("%d/%m/%Y")
        if isinstance(val, date):
            return val.strftime("%d/%m/%Y")
        s = str(val).strip()
        if len(s) >= 10:
            if hora and len(s) >= 16:
                return datetime.strptime(s[:16], "%Y-%m-%d %H:%M").strftime("%d/%m/%Y - %H:%M")
            return date.fromisoformat(s[:10]).strftime("%d/%m/%Y")
        return s
    except Exception:
        return str(val)

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from dashboard import auth, auth_repository
from utils.logger import get_logger
from dashboard.dashboard_repository import (
    buscar_ultima_execucao,
    buscar_total_processos,
    buscar_processos_monitorados,
    buscar_processos_sem_robo,
    buscar_total_orgaos,
    buscar_processos_por_status,
    buscar_ranking_orgaos,
    buscar_ultimas_movimentacoes,
    buscar_orgaos_sem_robo,
    buscar_ultimas_consultas,
    buscar_total_movimentacoes_recentes,
    buscar_movimentacoes_hoje_por_orgao,
    buscar_detalhe_movimentacoes_hoje,
    buscar_movimentacoes_do_dia_agrupadas,
    buscar_ultimas_movimentacoes_todos_processos,
    buscar_historico_7_dias,
    buscar_processo_por_id_dashboard,
    buscar_movimentacoes_do_processo,
    buscar_historico_consultas_do_processo,
    buscar_movimentacoes_do_mes,
    buscar_filtros_relatorio,
    buscar_dados_relatorio,
    buscar_todos_processos,
    buscar_filtros_processos,
)

from dashboard.dashboard_html import gerar_linhas_tabela

HOST = os.getenv("DASHBOARD_HOST", "localhost")
PORTA = int(os.getenv("DASHBOARD_PORT", "8000"))

_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_STATUS_FILE = os.path.join(_BASE_DIR, ".monitor_status.json")
_SCRIPT_MANUAL = os.path.join(_BASE_DIR, "services", "executar_monitoramento_manual.py")


def _ler_status_monitoramento():
    try:
        if os.path.exists(_STATUS_FILE):
            with open(_STATUS_FILE) as f:
                status = json.load(f)
            # Se marcado como running mas o processo não existe mais, auto-reseta
            if status.get("running"):
                pid = status.get("pid")
                processo_vivo = False
                if pid:
                    try:
                        os.kill(pid, 0)  # sinal 0 = só testa existência
                        processo_vivo = True
                    except (OSError, ProcessLookupError):
                        processo_vivo = False
                if not processo_vivo:
                    status["running"] = False
                    status["orgao_atual"] = ""
                    try:
                        with open(_STATUS_FILE, "w") as f:
                            json.dump(status, f)
                    except Exception:
                        pass
            return status
    except Exception:
        pass
    return {"running": False}

# ─────────────────────────────────────────────
# CSS compartilhado
# ─────────────────────────────────────────────
CSS_BASE = """
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    :root {
        --bg-page:     #f0f2f5;
        --bg-card:     #ffffff;
        --bg-section:  #ffffff;
        --bg-header:   #0f172a;
        --bg-th:       #f8fafc;
        --bg-hover:    #f1f5f9;
        --text-1:      #0f172a;
        --text-2:      #64748b;
        --text-3:      #94a3b8;
        --border:      #e2e8f0;
        --blue:        #2563eb;
        --green:       #16a34a;
        --orange:      #ea580c;
        --gray:        #64748b;
        --shadow-sm:   0 1px 3px rgba(0,0,0,.07), 0 1px 2px rgba(0,0,0,.04);
        --shadow-md:   0 4px 16px rgba(0,0,0,.10), 0 2px 4px rgba(0,0,0,.06);
        --shadow-lg:   0 8px 28px rgba(0,0,0,.14), 0 4px 8px rgba(0,0,0,.08);
        --radius:      12px;
        --radius-sm:   8px;
    }

    [data-theme="dark"] {
        --bg-page:     #0d1117;
        --bg-card:     #161b22;
        --bg-section:  #161b22;
        --bg-header:   #010409;
        --bg-th:       #1c2128;
        --bg-hover:    #1c2128;
        --text-1:      #e6edf3;
        --text-2:      #8b949e;
        --text-3:      #6e7681;
        --border:      #30363d;
        --blue:        #388bfd;
        --green:       #3fb950;
        --orange:      #fb8f44;
        --gray:        #8b949e;
        --shadow-sm:   0 1px 3px rgba(0,0,0,.30), 0 1px 2px rgba(0,0,0,.20);
        --shadow-md:   0 4px 16px rgba(0,0,0,.40), 0 2px 4px rgba(0,0,0,.30);
        --shadow-lg:   0 8px 28px rgba(0,0,0,.55), 0 4px 8px rgba(0,0,0,.40);
    }

    html { scroll-behavior: smooth; }
    body {
        font-family: 'Inter', Arial, sans-serif;
        background: var(--bg-page);
        color: var(--text-1);
        min-height: 100vh;
        font-size: 14px;
        line-height: 1.5;
        transition: background .25s, color .25s;
    }

    /* ── Topbar ── */
    .topbar {
        background: var(--bg-header);
        height: 58px;
        padding: 0 28px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        position: sticky;
        top: 0;
        z-index: 100;
        border-bottom: 1px solid rgba(255,255,255,.06);
    }
    .topbar-brand {
        display: flex;
        align-items: center;
        gap: 11px;
        text-decoration: none;
    }
    .topbar-icon {
        width: 34px; height: 34px;
        background: linear-gradient(135deg, #2563eb 0%, #7c3aed 100%);
        border-radius: 8px;
        display: flex; align-items: center; justify-content: center;
        font-size: 17px; flex-shrink: 0;
    }
    .topbar-name {
        font-size: 14px; font-weight: 700;
        color: #ffffff; letter-spacing: -.2px;
        line-height: 1.1;
    }
    .topbar-sub {
        font-size: 10px; color: rgba(255,255,255,.4);
        font-weight: 400; letter-spacing: .3px;
    }
    .topbar-right { display: flex; align-items: center; gap: 6px; }
    .topbar-nav { display: flex; gap: 2px; margin-right: 10px; }
    .topbar-nav a {
        padding: 6px 13px;
        border-radius: var(--radius-sm);
        font-size: 13px; font-weight: 500;
        color: rgba(255,255,255,.65);
        text-decoration: none;
        transition: background .15s, color .15s;
        white-space: nowrap;
    }
    .topbar-nav a:hover { background: rgba(255,255,255,.08); color: #fff; }
    .topbar-nav a.ativo { background: rgba(255,255,255,.11); color: #fff; }
    .btn-theme {
        width: 34px; height: 34px;
        border-radius: var(--radius-sm);
        border: 1px solid rgba(255,255,255,.12);
        background: rgba(255,255,255,.06);
        color: rgba(255,255,255,.75);
        cursor: pointer; font-size: 15px;
        display: flex; align-items: center; justify-content: center;
        transition: background .15s;
        flex-shrink: 0;
    }
    .btn-theme:hover { background: rgba(255,255,255,.13); color: #fff; }

    /* ── Layout ── */
    .page-content { max-width: 1300px; margin: 0 auto; padding: 28px 24px; }
    .page-header { margin-bottom: 22px; }
    .page-header h1 {
        font-size: 22px; font-weight: 700;
        color: var(--text-1); letter-spacing: -.4px;
    }
    .subtitulo { color: var(--text-2); font-size: 13px; margin-top: 3px; }

    /* ── Cards ── */
    .cards {
        display: grid;
        grid-template-columns: repeat(4, minmax(180px, 1fr));
        gap: 16px;
        margin-bottom: 22px;
    }
    .card {
        background: var(--bg-card);
        padding: 20px 22px;
        border-radius: var(--radius);
        box-shadow: var(--shadow-sm);
        border: 1px solid var(--border);
        border-top: 3px solid var(--blue);
        text-decoration: none;
        color: inherit;
        display: block;
        transition: box-shadow .15s, transform .15s;
    }
    .card:hover { box-shadow: var(--shadow-lg); transform: translateY(-2px); }
    .card.alerta  { border-top-color: var(--orange); }
    .card.sucesso { border-top-color: var(--green); }
    .card.neutro  { border-top-color: var(--gray); }
    .card h2 { font-size: 36px; font-weight: 800; color: var(--text-1); letter-spacing: -1.5px; line-height: 1; margin-bottom: 4px; }
    .card p  { color: var(--text-2); font-size: 13px; margin-top: 2px; }
    .card .hint { font-size: 11px; color: var(--text-3); margin-top: 6px; }

    /* ── Grid ── */
    .grid-duplo {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 20px;
        margin-bottom: 20px;
    }

    /* ── Section ── */
    section {
        background: var(--bg-section);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        padding: 22px 24px;
        margin-bottom: 20px;
        box-shadow: var(--shadow-sm);
        overflow-x: auto;
    }
    section h2 {
        font-size: 15px; font-weight: 700;
        color: var(--text-1); letter-spacing: -.3px;
        margin-bottom: 14px;
    }

    /* ── Tabelas ── */
    table { width: 100%; border-collapse: collapse; }
    th, td {
        padding: 10px 12px;
        text-align: left;
        font-size: 13px;
        vertical-align: middle;
        max-width: 300px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        border-bottom: 1px solid var(--border);
    }
    th {
        background: var(--bg-th);
        color: var(--text-2);
        font-size: 11px;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: .7px;
        border-bottom: 2px solid var(--border);
    }
    tr:hover td { background: var(--bg-hover); }
    tbody tr:last-child td { border-bottom: none; }

    /* ── Badges ── */
    .badge {
        display: inline-block;
        padding: 2px 9px;
        border-radius: 999px;
        font-size: 11px;
        font-weight: 600;
        letter-spacing: .2px;
    }
    .badge.verde   { background: #dcfce7; color: #15803d; }
    .badge.cinza   { background: #f1f5f9; color: #475569; }
    .badge.azul    { background: #dbeafe; color: #1d4ed8; }
    .badge.laranja { background: #ffedd5; color: #c2410c; }
    .badge.roxo    { background: #ede9fe; color: #6d28d9; }
    [data-theme="dark"] .badge.verde   { background: #0a2d1a; color: #4ade80; }
    [data-theme="dark"] .badge.cinza   { background: #1c2128; color: #8b949e; }
    [data-theme="dark"] .badge.azul    { background: #0c1f3d; color: #79b8ff; }
    [data-theme="dark"] .badge.laranja { background: #2e1200; color: #fb8f44; }
    [data-theme="dark"] .badge.roxo    { background: #1a0040; color: #c084fc; }

    /* ── Barra ── */
    .barra-container { background: var(--border); border-radius: 4px; height: 6px; width: 100%; }
    .barra { background: var(--orange); height: 6px; border-radius: 4px; }

    /* ── Calendário ── */
    .cal-semana { display: flex; gap: 10px; flex-wrap: wrap; }
    .cal-dia {
        display: flex; flex-direction: column; align-items: center;
        background: var(--bg-card);
        border: 1px solid var(--border);
        border-radius: 10px;
        padding: 12px 16px;
        box-shadow: var(--shadow-sm);
        min-width: 82px;
        text-decoration: none;
        color: inherit;
        transition: border-color .15s, box-shadow .15s, transform .15s;
    }
    .cal-dia:hover { border-color: var(--orange); box-shadow: var(--shadow-md); transform: translateY(-2px); }
    .cal-dia.hoje  { border-color: var(--blue); box-shadow: 0 0 0 2px rgba(37,99,235,.18); }
    .cal-dia.sem-dados { opacity: .5; }
    .cal-dia .dia-semana { font-size: 10px; font-weight: 700; color: var(--text-3); text-transform: uppercase; letter-spacing: .8px; }
    .cal-dia .dia-num    { font-size: 22px; font-weight: 800; color: var(--text-1); line-height: 1.2; }
    .cal-dia .dia-total  { font-size: 20px; font-weight: 800; color: var(--orange); margin-top: 4px; }
    .cal-dia .dia-label  { font-size: 10px; color: var(--text-3); }
    .cal-dia .bolinha    { width: 6px; height: 6px; border-radius: 50%; margin-top: 6px; background: var(--border); }
    .cal-dia.com-dados .bolinha { background: var(--orange); }

    /* ── Nav datas ── */
    .nav-data {
        display: flex; align-items: center; gap: 10px;
        margin-bottom: 20px; flex-wrap: wrap;
    }
    .nav-data a, .nav-data button {
        padding: 7px 14px;
        border-radius: var(--radius-sm);
        border: 1px solid var(--border);
        background: var(--bg-card);
        color: var(--text-1);
        cursor: pointer; font-size: 13px;
        text-decoration: none;
        transition: background .1s, border-color .1s;
        font-family: inherit;
    }
    .nav-data a:hover, .nav-data button:hover { background: var(--bg-hover); border-color: var(--blue); }
    .nav-data .data-atual { font-size: 16px; font-weight: 700; color: var(--text-1); }
    .nav-data input[type=date] {
        padding: 7px 12px;
        border-radius: var(--radius-sm);
        border: 1px solid var(--border);
        background: var(--bg-card);
        color: var(--text-1);
        font-size: 13px; cursor: pointer; font-family: inherit;
    }

    /* ── Misc ── */
    .vazio { text-align: center; color: var(--text-3); padding: 24px; font-size: 13px; }
    .footer { text-align: center; color: var(--text-3); margin-top: 40px; padding-bottom: 28px; font-size: 12px; }

    .btn {
        display: inline-flex; align-items: center; gap: 6px;
        padding: 8px 18px; margin-bottom: 12px;
        background: var(--blue);
        color: white; border: none;
        border-radius: var(--radius-sm);
        cursor: pointer; text-decoration: none;
        font-size: 13px; font-weight: 600;
        transition: opacity .15s, transform .1s;
        font-family: inherit;
    }
    .btn:hover { opacity: .88; transform: translateY(-1px); }
    .btn.secundario {
        background: var(--bg-card);
        color: var(--text-1);
        border: 1px solid var(--border);
    }
    .btn.secundario:hover { background: var(--bg-hover); }

    @media (max-width: 1000px) {
        .cards { grid-template-columns: repeat(2, 1fr); }
        .grid-duplo { grid-template-columns: 1fr; }
    }
    @media (max-width: 600px) {
        .cards { grid-template-columns: 1fr; }
        .page-content { padding: 16px; }
        .topbar { padding: 0 16px; }
        .topbar-nav { display: none; }
    }
"""

# JS injetado no <head> para evitar flash de tema errado ao carregar
_JS_THEME_INIT = """
<script>
(function(){
    var t = localStorage.getItem('ssa-theme') || 'light';
    document.documentElement.setAttribute('data-theme', t);
})();
</script>"""

# JS de toggle de tema (injetado no <body>)
_JS_THEME_TOGGLE = """
<script>
function _ssaToggleTheme() {
    var cur  = document.documentElement.getAttribute('data-theme');
    var next = cur === 'dark' ? 'light' : 'dark';
    document.documentElement.setAttribute('data-theme', next);
    localStorage.setItem('ssa-theme', next);
    document.getElementById('btn-theme').textContent = next === 'dark' ? '☀' : '🌙';
}
// Corrige ícone ao carregar
(function(){
    var t = document.documentElement.getAttribute('data-theme');
    var b = document.getElementById('btn-theme');
    if (b) b.textContent = t === 'dark' ? '☀' : '🌙';
})();
</script>"""


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def _badge_status(status):
    s = (status or "").lower()
    if "finalizado" in s or "deferido" in s or "encerrado" in s:
        return f'<span class="badge verde">{escape(status or "")}</span>'
    if "análise" in s or "andamento" in s:
        return f'<span class="badge azul">{escape(status or "")}</span>'
    if "indeferido" in s:
        return f'<span class="badge laranja">{escape(status or "")}</span>'
    return f'<span class="badge cinza">{escape(status or "Sem status")}</span>'


def _topbar(pagina_atual="/", usuario=None):
    links = [
        ("/", "Dashboard"),
        ("/calendario", "Calendário"),
        ("/movimentacoes-hoje", "Movimentações Hoje"),
        ("/processos", "Processos"),
        ("/relatorio", "Relatório"),
    ]
    if usuario and usuario.get("is_admin"):
        links.append(("/admin/usuarios", "Usuários"))

    nav_html = ""
    for href, label in links:
        cls = "ativo" if href == pagina_atual else ""
        nav_html += f'<a href="{href}" class="{cls}">{escape(label)}</a>'

    usuario_html = ""
    if usuario:
        nome = escape(str(usuario.get("nome") or usuario.get("email") or ""))
        usuario_html = (
            f'<span style="font-size:12px;color:rgba(255,255,255,.55);margin-right:4px;">{nome}</span>'
            f'<a href="/logout" class="btn-theme" style="width:auto;padding:0 12px;text-decoration:none;font-size:12px;">Sair</a>'
        )

    return f"""
    <header class="topbar">
        <a href="/" class="topbar-brand">
            <div class="topbar-icon">&#128202;</div>
            <div>
                <div class="topbar-name">SSA Monitor</div>
                <div class="topbar-sub">PROCESSOS</div>
            </div>
        </a>
        <div class="topbar-right">
            <nav class="topbar-nav">{nav_html}</nav>
            {usuario_html}
            <button class="btn-theme" id="btn-theme" onclick="_ssaToggleTheme()" title="Alternar tema">🌙</button>
        </div>
    </header>"""


# ─────────────────────────────────────────────
# PÁGINA PRINCIPAL — Dashboard
# ─────────────────────────────────────────────
def _gerar_calendario_7_dias(historico):
    """Gera os 7 cards do mini-calendário com links para cada dia."""
    # Monta dict data → dados para lookup rápido
    por_dia = {}
    for row in historico:
        d = row.get("dia")
        if d:
            por_dia[str(d)] = row

    hoje = date.today()
    dias_semana_pt = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    html = '<div class="cal-semana">'

    for i in range(6, -1, -1):
        dia = hoje - timedelta(days=i)
        chave = str(dia)
        dados = por_dia.get(chave, {})
        total = dados.get("total_processos", 0)
        nome_dia = dias_semana_pt[dia.weekday()]
        is_hoje = dia == hoje
        tem_dados = total > 0

        cls = "cal-dia"
        if is_hoje:
            cls += " hoje"
        if tem_dados:
            cls += " com-dados"
        else:
            cls += " sem-dados"

        link = f"/movimentacoes-hoje?data={chave}"
        label_dia = "hoje" if is_hoje else dia.strftime("%d/%m")

        html += f"""
        <a href="{link}" class="{cls}" title="Ver {label_dia}">
            <span class="dia-semana">{nome_dia}</span>
            <span class="dia-num">{dia.day:02d}</span>
            <span class="dia-total">{total if tem_dados else '—'}</span>
            <span class="dia-label">{'processo' + ('s' if total != 1 else '') if tem_dados else 'sem mov.'}</span>
            <span class="bolinha"></span>
        </a>"""

    html += "</div>"
    return html


def gerar_html_dashboard(usuario=None):
    total_processos       = buscar_total_processos()
    processos_monitorados = buscar_processos_monitorados()
    total_orgaos          = buscar_total_orgaos()
    novas_movimentacoes   = buscar_total_movimentacoes_recentes()
    processos_por_status  = buscar_processos_por_status()
    ranking_orgaos        = buscar_ranking_orgaos()
    ultimas_movimentacoes = buscar_ultimas_movimentacoes()
    orgaos_sem_robo       = buscar_orgaos_sem_robo()
    ultimas_consultas     = buscar_ultimas_consultas()
    mov_por_orgao         = buscar_movimentacoes_hoje_por_orgao()
    historico_7           = buscar_historico_7_dias()
    ultima_execucao       = buscar_ultima_execucao()

    # Linhas da tabela de status com badges
    html_status = ""
    for item in processos_por_status:
        s = escape(str(item.get("status") or ""))
        t = item.get("total", 0)
        html_status += f"<tr><td>{_badge_status(s)}</td><td><strong>{t}</strong></td></tr>"
    if not html_status:
        html_status = '<tr><td colspan="2" class="vazio">Sem dados</td></tr>'

    html_ranking = gerar_linhas_tabela(ranking_orgaos, ["orgao", "total_processos"])

    html_movimentacoes = ""
    for m in ultimas_movimentacoes:
        html_movimentacoes += (
            f"<tr>"
            f"<td>{escape(str(m.get('numero_processo') or ''))}</td>"
            f"<td>{escape(str(m.get('empresa') or ''))}</td>"
            f"<td>{escape(str(m.get('orgao') or ''))}</td>"
            f"<td>{escape(_fmt_data(m.get('data_movimento')))}</td>"
            f"<td>{escape(str(m.get('descricao') or ''))}</td>"
            f"</tr>"
        )
    if not html_movimentacoes:
        html_movimentacoes = '<tr><td colspan="5" class="vazio">Nenhuma movimentação recente</td></tr>'

    html_orgaos_sem_robo = gerar_linhas_tabela(orgaos_sem_robo, ["nome", "total_processos", "url"])

    html_consultas = ""
    for c in ultimas_consultas:
        sc = str(c.get("status_consulta") or "")
        badge = _badge_status(sc) if sc == "OK" else f'<span class="badge laranja">{escape(sc)}</span>'
        html_consultas += (
            f"<tr>"
            f"<td>{escape(str(c.get('numero_processo') or ''))}</td>"
            f"<td>{escape(str(c.get('empresa') or ''))}</td>"
            f"<td>{escape(str(c.get('orgao') or ''))}</td>"
            f"<td>{badge}</td>"
            f"<td>{escape(_fmt_data(c.get('data_consulta'), hora=True))}</td>"
            f"</tr>"
        )
    if not html_consultas:
        html_consultas = '<tr><td colspan="5" class="vazio">Nenhuma consulta registrada</td></tr>'

    # Mini breakdown de movimentações por prefeitura no card
    mini_lista = ""
    if mov_por_orgao:
        for item in mov_por_orgao[:4]:
            orgao = escape(str(item.get("orgao") or ""))
            total = item.get("total_processos", 0)
            mini_lista += f'<div style="font-size:12px;color:#6b7280;margin-top:3px;">• {orgao}: <strong>{total}</strong></div>'
    else:
        mini_lista = '<div style="font-size:12px;color:#9ca3af;margin-top:4px;">Nenhuma hoje</div>'

    html_calendario = _gerar_calendario_7_dias(historico_7)

    ultima_str = _fmt_data(ultima_execucao, hora=True) if ultima_execucao else "—"

    is_admin = usuario and usuario.get("is_admin")
    status_mon = _ler_status_monitoramento() if is_admin else {}
    running = status_mon.get("running", False)
    btn_label = "⏳ Executando..." if running else "▶ Executar agora"
    btn_disabled = "disabled" if running else ""
    concluidos_init = status_mon.get("concluidos", 0)
    total_init = status_mon.get("total", 0)
    orgao_init = status_mon.get("orgao_atual", "")
    pct_init = round(concluidos_init / total_init * 100) if total_init else 0

    btn_executar_html = ""
    js_admin = ""
    if is_admin:
        btn_executar_html = f"""
        <div style="text-align:right;">
            <button id="btn-executar" onclick="executarMonitoramento()" {btn_disabled}
                    style="background:var(--blue);color:#fff;border:none;border-radius:8px;
                           padding:10px 20px;font-size:14px;font-weight:600;cursor:pointer;
                           opacity:{0.6 if running else 1};font-family:inherit;">
                {btn_label}
            </button>
            <div id="progresso-wrap" style="display:{'block' if running else 'none'};
                 margin-top:10px;padding:12px 14px;background:var(--bg-card);
                 border:1px solid var(--border);border-radius:10px;text-align:left;">
                <div style="display:flex;justify-content:space-between;
                            font-size:12px;color:var(--text-2);margin-bottom:6px;">
                    <span id="progresso-label">{'Iniciando...' if not total_init else f'{concluidos_init} / {total_init} processos'}</span>
                    <span id="progresso-pct" style="font-weight:700;">{f'{pct_init}%' if total_init else ''}</span>
                </div>
                <div style="background:var(--border);border-radius:4px;height:8px;overflow:hidden;">
                    <div id="progresso-barra"
                         style="background:var(--blue);height:8px;border-radius:4px;
                                width:{pct_init}%;transition:width .6s ease;"></div>
                </div>
                <div id="progresso-orgao"
                     style="font-size:11px;color:var(--text-3);margin-top:5px;">
                    {'&#128205; ' + orgao_init if orgao_init else ''}
                </div>
            </div>
        </div>"""
        js_admin = """
<script>
function _atualizarBarra(concluidos, total, orgao) {
    var pct = total > 0 ? Math.round(concluidos / total * 100) : 0;
    document.getElementById('progresso-label').textContent =
        total > 0 ? concluidos + ' / ' + total + ' processos' : 'Iniciando...';
    document.getElementById('progresso-pct').textContent =
        total > 0 ? pct + '%' : '';
    document.getElementById('progresso-barra').style.width = pct + '%';
    document.getElementById('progresso-orgao').textContent =
        orgao ? '📍 ' + orgao : '';
}
function verificarStatus() {
    fetch('/api/admin/status-monitoramento')
        .then(r => r.json())
        .then(data => {
            var btn  = document.getElementById('btn-executar');
            var span = document.getElementById('ultima-execucao');
            var wrap = document.getElementById('progresso-wrap');
            if (data.running) {
                btn.textContent   = '⏳ Executando...';
                btn.disabled      = true;
                btn.style.opacity = '0.6';
                wrap.style.display = 'block';
                _atualizarBarra(data.concluidos || 0, data.total || 0, data.orgao_atual || '');
                setTimeout(verificarStatus, 2000);
            } else {
                btn.textContent   = '▶ Executar agora';
                btn.disabled      = false;
                btn.style.opacity = '1';
                wrap.style.display = 'none';
                if (data.ultima_execucao) {
                    span.textContent = 'Última execução: ' + data.ultima_execucao;
                }
            }
        })
        .catch(function(){});
}
function executarMonitoramento() {
    fetch('/api/admin/executar-monitoramento', {method: 'POST'})
        .then(r => r.json())
        .then(data => {
            if (data.ok) {
                var btn  = document.getElementById('btn-executar');
                var wrap = document.getElementById('progresso-wrap');
                btn.textContent   = '⏳ Executando...';
                btn.disabled      = true;
                btn.style.opacity = '0.6';
                wrap.style.display = 'block';
                _atualizarBarra(0, 0, '');
                setTimeout(verificarStatus, 2000);
            }
        })
        .catch(function(){});
}
document.addEventListener('DOMContentLoaded', function() {
    if (""" + ("true" if running else "false") + """) { verificarStatus(); }
});
</script>"""

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta http-equiv="refresh" content="30">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>SSA Monitor Processos</title>
    {_JS_THEME_INIT}
    <style>{CSS_BASE}</style>
</head>
<body>
    {_topbar("/", usuario)}
    <div class="page-content">
        <div class="page-header" style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:12px;">
            <div>
                <h1>Dashboard</h1>
                <div class="subtitulo">Visão geral · atualiza automaticamente a cada 30s · <span id="ultima-execucao">Última execução: {ultima_str}</span></div>
            </div>
            {btn_executar_html}
        </div>

        <div class="cards">
            <div class="card">
                <h2>{total_processos}</h2>
                <p>Total de processos</p>
            </div>
            <div class="card sucesso">
                <h2>{processos_monitorados}</h2>
                <p>Processos monitorados</p>
            </div>
            <a href="/movimentacoes-hoje" class="card alerta">
                <h2>{novas_movimentacoes}</h2>
                <p>Processos com mov. hoje</p>
                {mini_lista}
                <div class="hint">Clique para ver detalhes ›</div>
            </a>
            <div class="card neutro">
                <h2>{total_orgaos}</h2>
                <p>Total de órgãos/links</p>
            </div>
        </div>

        <section>
            <h2>Histórico dos últimos 7 dias</h2>
            <p style="font-size:12px;color:var(--text-2);margin-top:-8px;margin-bottom:16px;">
                Clique em qualquer dia para ver os detalhes dos processos com movimentação.
            </p>
            {html_calendario}
        </section>

        <div class="grid-duplo">
            <section>
                <h2>Processos por status</h2>
                <table>
                    <thead><tr><th>Status</th><th>Total</th></tr></thead>
                    <tbody>{html_status}</tbody>
                </table>
            </section>
            <section>
                <h2>Ranking por prefeitura</h2>
                <table>
                    <thead><tr><th>Prefeitura</th><th>Total de processos</th></tr></thead>
                    <tbody>{html_ranking}</tbody>
                </table>
            </section>
        </div>

        <section>
            <h2>Últimas movimentações registradas</h2>
            <table>
                <thead>
                    <tr>
                        <th>Processo</th><th>Empresa</th><th>Prefeitura</th>
                        <th>Data mov.</th><th>Descrição</th>
                    </tr>
                </thead>
                <tbody>{html_movimentacoes}</tbody>
            </table>
        </section>

        <section>
            <h2>Últimas consultas realizadas</h2>
            <table>
                <thead>
                    <tr>
                        <th>Processo</th><th>Empresa</th><th>Prefeitura</th>
                        <th>Resultado</th><th>Data/hora</th>
                    </tr>
                </thead>
                <tbody>{html_consultas}</tbody>
            </table>
        </section>

        <section>
            <h2>Prefeituras sem robô configurado</h2>
            <table>
                <thead><tr><th>Órgão</th><th>Processos</th><th>URL</th></tr></thead>
                <tbody>{html_orgaos_sem_robo}</tbody>
            </table>
        </section>

        <div class="footer">SSA Monitor Processos · Dashboard Web</div>
    </div>
    {_JS_THEME_TOGGLE}
    {js_admin}
</body>
</html>"""


# ─────────────────────────────────────────────
# PÁGINA: /movimentacoes-hoje  (aceita ?data=YYYY-MM-DD)
# ─────────────────────────────────────────────
def _nav_datas(data_selecionada: date) -> str:
    """Gera a barra de navegação de datas (← ontem | data atual | amanhã →)."""
    hoje       = date.today()
    anterior   = data_selecionada - timedelta(days=1)
    proximo    = data_selecionada + timedelta(days=1)
    dias_semana_pt = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
    nome_dia = dias_semana_pt[data_selecionada.weekday()]

    label_data = (
        f"Hoje — {data_selecionada.strftime('%d/%m/%Y')}"
        if data_selecionada == hoje
        else f"{nome_dia}, {data_selecionada.strftime('%d/%m/%Y')}"
    )

    proximo_html = (
        f'<a href="/movimentacoes-hoje?data={proximo}">Amanhã &rarr;</a>'
        if proximo <= hoje
        else '<span style="color:#d1d5db;padding:8px 16px;border-radius:8px;border:1px solid #e5e7eb;">Amanhã →</span>'
    )

    return f"""
    <div class="nav-data">
        <a href="/movimentacoes-hoje?data={anterior}">&larr; Dia anterior</a>
        <span class="data-atual">📅 {label_data}</span>
        {proximo_html}
        <form method="get" action="/movimentacoes-hoje" style="display:inline-flex;align-items:center;gap:8px;">
            <input type="date" name="data" value="{data_selecionada}"
                   max="{hoje}" onchange="this.form.submit()">
        </form>
        {'<a href="/movimentacoes-hoje">↩ Ir para hoje</a>' if data_selecionada != hoje else ''}
    </div>"""


def _html_movimentacoes_expandidas(movs):
    """Gera o conteúdo HTML da linha expandida com as movimentações do dia."""
    if not movs:
        return '<p style="color:#6b7280;margin:0">Nenhuma movimentação registrada para este dia.</p>'
    linhas = ""
    for m in movs:
        dt  = escape(_fmt_data(m.get("data_movimento")))
        hr  = escape(str(m.get("hora_captura") or "")[:5])
        desc = escape(str(m.get("descricao") or "").strip()[:300])
        # Só exibe linhas que parecem tramitações reais (têm data e texto)
        if not desc or len(desc) < 5:
            continue
        linhas += f"""
        <tr style="background:#f8fafc;">
            <td style="white-space:nowrap;color:#6b7280;font-size:12px;padding:6px 10px;">{dt}</td>
            <td style="color:#6b7280;font-size:12px;padding:6px 10px;">{hr}</td>
            <td style="font-size:13px;white-space:normal;max-width:600px;padding:6px 10px;">{desc}</td>
        </tr>"""
    if not linhas:
        return '<p style="color:#6b7280;margin:0">Detalhes não disponíveis para este dia.</p>'
    return f"""
    <table style="width:100%;border-collapse:collapse;margin-top:6px;">
        <thead>
            <tr>
                <th style="font-size:11px;color:#9ca3af;text-align:left;padding:4px 10px;white-space:nowrap;">Data mov.</th>
                <th style="font-size:11px;color:#9ca3af;text-align:left;padding:4px 10px;">Capturado às</th>
                <th style="font-size:11px;color:#9ca3af;text-align:left;padding:4px 10px;">Descrição</th>
            </tr>
        </thead>
        <tbody>{linhas}</tbody>
    </table>"""


def gerar_html_movimentacoes_hoje(data_str=None, usuario=None):
    hoje = date.today()
    try:
        data_sel = date.fromisoformat(data_str) if data_str else hoje
        if data_sel > hoje:
            data_sel = hoje
    except (ValueError, TypeError):
        data_sel = hoje

    data_param = str(data_sel) if data_sel != hoje else None

    processos     = buscar_detalhe_movimentacoes_hoje(data_param)
    mov_por_orgao = buscar_movimentacoes_hoje_por_orgao(data_param)
    total_dia     = buscar_total_movimentacoes_recentes(data_param)
    movs_agrupadas       = buscar_movimentacoes_do_dia_agrupadas(data_param)
    movs_historico       = buscar_ultimas_movimentacoes_todos_processos()

    # Cards de prefeitura — clicáveis para filtrar a tabela
    cards_orgaos = ""
    for item in mov_por_orgao:
        orgao_raw = str(item.get("orgao") or "")
        orgao     = escape(orgao_raw)
        tp = item.get("total_processos", 0)
        tm = item.get("total_movimentacoes", 0)
        cards_orgaos += f"""
        <div class="card alerta" style="min-width:160px;cursor:pointer;"
             onclick="filtrarPrefeitura('{orgao_raw}')"
             title="Filtrar por {orgao}">
            <h2>{tp}</h2>
            <p>{orgao}</p>
            <div class="hint">{tm} movimentaç{'ão' if tm == 1 else 'ões'} · clique para filtrar</div>
        </div>"""

    if not cards_orgaos:
        cards_orgaos = '<p style="color:#6b7280">Nenhuma movimentação detectada neste dia.</p>'

    STATUS_CONCLUIDOS = {"Deferido", "Finalizado", "Indeferido", "Encerrado"}

    def _gerar_linha(p, concluido=False):
        pid       = p.get("processo_id", "")
        num       = escape(str(p.get("numero_processo") or ""))
        emp       = escape(str(p.get("empresa") or "—"))
        orgao_raw = str(p.get("orgao") or "")
        orgao     = escape(orgao_raw)
        status_val= str(p.get("status_atual") or "")
        status    = _badge_status(status_val)
        status_raw= escape(status_val)
        total     = p.get("total_movimentacoes_hoje", 0)
        dt_mov    = escape(_fmt_data(p.get("data_ultimo_movimento")))
        ult_c     = escape(_fmt_data(p.get("ultima_consulta"), hora=True))

        movs_hoje_proc = movs_agrupadas.get(pid, [])
        movs_hist_proc = movs_historico.get(pid, [])

        if total > 0:
            indicador = f'<span class="badge verde">✔ {total} nova{"s" if total > 1 else ""}</span>'
            cor_linha = "linha-destaque"
        elif movs_hist_proc:
            indicador = '<span class="badge cinza">— sem mov. hoje</span>'
            cor_linha = ""
        else:
            indicador = '<span class="badge cinza">— sem mov.</span>'
            cor_linha = ""

        opacidade = "opacity:0.55;" if concluido else ""
        link_detalhe = f"/processo/{pid}"

        return f"""
        <tr data-prefeitura="{escape(orgao_raw)}"
            data-empresa="{emp}"
            data-status="{status_raw}"
            data-tem-mov="{1 if total > 0 else 0}"
            data-concluido="{'1' if concluido else '0'}"
            style="{opacidade}"
            onclick="window.location='{link_detalhe}'"
            title="Ver detalhes e movimentacoes"
            class="linha-processo{' concluido' if concluido else ''}{' ' + cor_linha if cor_linha else ''}">
            <td>
                <a href="{link_detalhe}" style="color:inherit;text-decoration:none;font-weight:700;">
                    {num}
                </a>
            </td>
            <td style="max-width:160px">{emp}</td>
            <td>{orgao}</td>
            <td>{status}</td>
            <td>{indicador}</td>
            <td>{dt_mov}</td>
            <td style="color:#9ca3af;font-size:12px">{ult_c} <span style="float:right">›</span></td>
        </tr>"""

    ativos     = [p for p in processos if str(p.get("status_atual") or "") not in STATUS_CONCLUIDOS]
    concluidos = [p for p in processos if str(p.get("status_atual") or "") in STATUS_CONCLUIDOS]

    linhas = "".join(_gerar_linha(p, concluido=False) for p in ativos)

    if concluidos:
        linhas += f"""
        <tr class="separador-concluidos" data-prefeitura="" data-empresa="" data-status="" data-tem-mov="0" data-concluido="1">
            <td colspan="7" style="
                text-align:center;font-size:11px;font-weight:600;letter-spacing:.08em;
                text-transform:uppercase;color:var(--text-3);padding:10px 0 6px;
                border-top:2px dashed var(--border);background:transparent;pointer-events:none;">
                Processos concluídos ({len(concluidos)})
            </td>
        </tr>"""
        linhas += "".join(_gerar_linha(p, concluido=True) for p in concluidos)

    if not linhas:
        linhas = '<tr><td colspan="7" class="vazio">Nenhum processo encontrado.</td></tr>'

    titulo = (
        "Movimentações de Hoje"
        if data_sel == hoje
        else f"Movimentações de {data_sel.strftime('%d/%m/%Y')}"
    )

    js = """
    <script>
    function popularSelect(id, attr) {
        const vals = new Set();
        document.querySelectorAll('tr[data-prefeitura]').forEach(r => {
            const v = r.dataset[attr];
            if (v) vals.add(v);
        });
        const sel = document.getElementById(id);
        [...vals].sort().forEach(v => {
            const opt = document.createElement('option');
            opt.value = v; opt.textContent = v;
            sel.appendChild(opt);
        });
    }

    function aplicarFiltros() {
        const texto   = document.getElementById('filtro-texto').value.toLowerCase();
        const pref    = document.getElementById('filtro-prefeitura').value;
        const emp     = document.getElementById('filtro-empresa').value;
        const status  = document.getElementById('filtro-status').value;

        let visiveis = 0;
        document.querySelectorAll('tr[data-prefeitura]').forEach(r => {
            const ok =
                (!pref   || r.dataset.prefeitura === pref) &&
                (!emp    || r.dataset.empresa    === emp)  &&
                (!status || r.dataset.status     === status) &&
                (!texto  || r.cells[0].textContent.toLowerCase().includes(texto)
                         || r.cells[1].textContent.toLowerCase().includes(texto));
            r.style.display = ok ? '' : 'none';
            if (ok) visiveis++;
        });

        const temFiltro = pref || emp || status || texto;
        const av = document.getElementById('aviso-filtro');
        if (temFiltro) {
            av.style.display = '';
            av.innerHTML = 'Exibindo <strong>' + visiveis + '</strong> processo(s) &nbsp;'
                + '<a href="#" onclick="limparFiltros();return false;" style="color:#2563eb;">limpar filtros ✕</a>';
        } else {
            av.style.display = 'none';
        }
    }

    function limparFiltros() {
        document.getElementById('filtro-texto').value = '';
        document.getElementById('filtro-prefeitura').value = '';
        document.getElementById('filtro-empresa').value = '';
        document.getElementById('filtro-status').value = '';
        aplicarFiltros();
    }

    function filtrarPrefeitura(nome) {
        document.getElementById('filtro-prefeitura').value = nome;
        aplicarFiltros();
        document.getElementById('barra-filtros').scrollIntoView({behavior:'smooth', block:'nearest'});
    }

    document.addEventListener('DOMContentLoaded', () => {
        popularSelect('filtro-prefeitura', 'prefeitura');
        popularSelect('filtro-empresa',    'empresa');
        popularSelect('filtro-status',     'status');
    });
    </script>
    """

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{escape(titulo)} · SSA Monitor</title>
    {_JS_THEME_INIT}
    <style>{CSS_BASE}
    tr.linha-processo {{ cursor: pointer; }}
    tr.linha-processo:hover td {{ background: var(--bg-hover) !important; }}
    tr.linha-destaque td {{ background: #fff8f0; color: var(--text-main); }}
    [data-theme="dark"] tr.linha-destaque td {{ background: #2a1800; color: var(--text-main); }}
    tr.linha-destaque:hover td {{ background: #ffe8d6 !important; }}
    [data-theme="dark"] tr.linha-destaque:hover td {{ background: #3a2200 !important; }}
    #aviso-filtro {{ background: #dbeafe; color: #1e40af; }}
    [data-theme="dark"] #aviso-filtro {{ background: #0c1f3d; color: #79b8ff; }}
    </style>
</head>
<body>
    {_topbar("/movimentacoes-hoje", usuario)}
    <div class="page-content">
        <div class="page-header">
            <h1>{escape(titulo)}</h1>
            <div class="subtitulo">Todos os processos ativos · movimentações detectadas pelo robô neste dia</div>
        </div>

        {_nav_datas(data_sel)}

        <div style="display:flex;gap:14px;flex-wrap:wrap;margin-bottom:22px;">
            <div class="card alerta" style="min-width:160px;">
                <h2>{total_dia}</h2>
                <p>Processos com mov. neste dia</p>
            </div>
            {cards_orgaos}
        </div>

        <section>
            <h2>Situação de todos os processos</h2>
            <div style="font-size:12px;color:var(--text-2);margin-bottom:10px;margin-top:-6px;">
                <span class="badge verde">✔ N novas</span> = detectadas neste dia &nbsp;|&nbsp;
                <span class="badge cinza">— sem mov. hoje</span> = sem novidade, mas tem histórico &nbsp;|&nbsp;
                <strong>Clique em qualquer linha</strong> para detalhes
            </div>
            <div id="barra-filtros" style="display:flex;flex-wrap:wrap;gap:10px;margin-bottom:14px;align-items:center;">
                <input id="filtro-texto" type="text" placeholder="Buscar processo ou empresa..."
                    oninput="aplicarFiltros()"
                    style="flex:1;min-width:200px;padding:8px 12px;border:1px solid var(--border);
                           border-radius:8px;font-size:13px;background:var(--bg-card);color:var(--text-1);">
                <select id="filtro-prefeitura" onchange="aplicarFiltros()"
                    style="padding:8px 12px;border:1px solid var(--border);border-radius:8px;
                           font-size:13px;background:var(--bg-card);color:var(--text-1);cursor:pointer;">
                    <option value="">Todas as prefeituras</option>
                </select>
                <select id="filtro-empresa" onchange="aplicarFiltros()"
                    style="padding:8px 12px;border:1px solid var(--border);border-radius:8px;
                           font-size:13px;background:var(--bg-card);color:var(--text-1);cursor:pointer;">
                    <option value="">Todas as empresas</option>
                </select>
                <select id="filtro-status" onchange="aplicarFiltros()"
                    style="padding:8px 12px;border:1px solid var(--border);border-radius:8px;
                           font-size:13px;background:var(--bg-card);color:var(--text-1);cursor:pointer;">
                    <option value="">Todos os status</option>
                </select>
            </div>
            <div id="aviso-filtro" style="display:none;
                 padding:8px 14px;border-radius:8px;margin-bottom:10px;font-size:13px;"></div>
            <table>
                <thead>
                    <tr>
                        <th>Processo</th>
                        <th>Empresa</th>
                        <th>Prefeitura</th>
                        <th>Status</th>
                        <th>Hoje</th>
                        <th>Último mov.</th>
                        <th>Última consulta</th>
                    </tr>
                </thead>
                <tbody>{linhas}</tbody>
            </table>
        </section>

        <div class="footer">SSA Monitor Processos · Movimentações Hoje</div>
    </div>
    {js}
    {_JS_THEME_TOGGLE}
</body>
</html>"""


# ─────────────────────────────────────────────
# PÁGINA: /calendario  — visão mensal (CSS Grid, estilo Google Agenda)
# ─────────────────────────────────────────────
def gerar_html_calendario(ano: int, mes: int, usuario=None):
    import calendar as _cal

    hoje = date.today()
    primeiro_dia = date(ano, mes, 1)

    if mes == 1:
        mes_ant, ano_ant = 12, ano - 1
    else:
        mes_ant, ano_ant = mes - 1, ano
    if mes == 12:
        mes_prox, ano_prox = 1, ano + 1
    else:
        mes_prox, ano_prox = mes + 1, ano

    MESES_PT = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    MESES_ABR = ["", "jan", "fev", "mar", "abr", "mai", "jun",
                 "jul", "ago", "set", "out", "nov", "dez"]

    rows = buscar_movimentacoes_do_mes(ano, mes)

    por_dia: dict = {}
    for r in rows:
        dia = r["dia"]
        if isinstance(dia, str):
            dia = date.fromisoformat(dia[:10])
        elif hasattr(dia, "date"):
            dia = dia.date()
        por_dia.setdefault(dia, []).append(r)

    # monthdatescalendar devolve objetos date reais (semana começa no domingo=6)
    semanas = _cal.Calendar(firstweekday=6).monthdatescalendar(ano, mes)

    DIAS_ABR = ["DOM.", "SEG.", "TER.", "QUA.", "QUI.", "SEX.", "SÁB."]

    # ── cabeçalho (7 divs de header) ──
    header_html = "".join(
        f'<div class="gcal-th">{d}</div>' for d in DIAS_ABR
    )

    # ── células de dias ──
    MAX_CHIPS = 3
    cells_html = ""
    for semana in semanas:
        for d in semana:
            is_mes_atual = d.month == mes
            is_hoje      = d == hoje
            movs         = por_dia.get(d, [])

            cls = "gcal-cell"
            if not is_mes_atual:
                cls += " gcal-outro"

            if d.day == 1 and not is_mes_atual:
                label_num = str(d.day) + " " + MESES_ABR[d.month]
            else:
                label_num = str(d.day)

            num_cls  = "gcal-num gcal-num-hoje" if is_hoje else "gcal-num"
            num_html = f'<div class="gcal-head"><span class="{num_cls}">{label_num}</span></div>'

            chips = ""
            for mv in movs[:MAX_CHIPS]:
                pid        = mv["processo_id"]
                num_p      = escape(str(mv["numero_processo"] or ""))
                emp        = escape(str(mv["empresa"] or ""))
                total      = mv["total"]
                chip_label = emp if emp and emp not in ("—", "") else num_p
                chips += (
                    f'<a href="/processo/{pid}" class="gcal-chip"'
                    f' title="{num_p} — {emp} ({total} mov.)">'
                    f'<span class="gcal-chip-label">{chip_label}</span>'
                    f'<span class="gcal-chip-n">{total}</span>'
                    f'</a>'
                )
            if len(movs) > MAX_CHIPS:
                extra     = len(movs) - MAX_CHIPS
                data_link = d.strftime("%Y-%m-%d")
                chips += (
                    f'<a href="/movimentacoes-hoje?data={data_link}"'
                    f' class="gcal-chip gcal-mais">+{extra} mais</a>'
                )

            cells_html += (
                f'<div class="{cls}">'
                f'{num_html}'
                f'<div class="gcal-events">{chips}</div>'
                f'</div>'
            )

    total_mes    = sum(len(v) for v in por_dia.values())
    dias_com_mov = len(por_dia)

    mes_ant_str  = str(mes_ant).zfill(2)
    mes_prox_str = str(mes_prox).zfill(2)
    pode_avancar = primeiro_dia < hoje.replace(day=1) or (ano == hoje.year and mes == hoje.month)
    if pode_avancar:
        nav_prox = (
            '<a class="nav-btn" href="/calendario?ano=' + str(ano_prox)
            + '&mes=' + mes_prox_str + '">' + MESES_PT[mes_prox] + ' &rarr;</a>'
        )
    else:
        nav_prox = (
            '<span class="nav-btn nav-btn-off">'
            + MESES_PT[mes_prox] + ' &rarr;</span>'
        )

    css_cal = """
/* ── Calendário (CSS Grid, estilo Google Agenda) ── */
.nav-mes {
    display: flex; align-items: center; gap: 16px;
    margin-bottom: 20px; flex-wrap: wrap;
}
.nav-mes .mes-atual { font-size: 22px; font-weight: 800; color: var(--text-1); }
.nav-btn {
    padding: 7px 18px; border: 1px solid var(--border); border-radius: 8px;
    color: var(--text-1); text-decoration: none; font-weight: 500;
    background: var(--bg-card); white-space: nowrap; display: inline-block;
}
.nav-btn:hover { background: var(--bg-hover); }
.nav-btn-off { color: var(--text-3); cursor: default; }

.gcal-wrap { overflow-x: auto; }
.gcal {
    display: grid;
    grid-template-columns: repeat(7, 1fr);
    gap: 1px;
    background: var(--border);
    border: 1px solid var(--border);
    border-radius: 12px;
    overflow: hidden;
    min-width: 700px;
}
.gcal-th {
    background: var(--bg-th);
    text-align: center;
    padding: 9px 4px;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: .05em;
    color: var(--text-2);
}
.gcal-cell {
    background: var(--bg-card);
    min-height: 110px;
    padding: 4px 5px 6px;
    overflow: hidden;
    display: flex;
    flex-direction: column;
}
.gcal-outro {
    background: var(--bg-page);
}
.gcal-outro .gcal-num {
    color: var(--text-3);
}
.gcal-head {
    display: flex;
    justify-content: center;
    margin-bottom: 3px;
}
.gcal-num {
    display: inline-block;
    width: 26px; height: 26px; line-height: 26px;
    text-align: center; border-radius: 50%;
    font-size: 12px; font-weight: 600;
    color: var(--text-2);
}
.gcal-num-hoje {
    background: var(--blue);
    color: #fff !important;
    font-weight: 700;
}
.gcal-events {
    display: flex; flex-direction: column; gap: 2px;
    flex: 1; overflow: hidden;
}
.gcal-chip {
    display: flex; align-items: center; gap: 4px;
    padding: 2px 5px; border-radius: 4px;
    background: #1a73e8; color: #fff;
    font-size: 11px; line-height: 1.35;
    text-decoration: none; overflow: hidden; white-space: nowrap;
    transition: filter .15s;
}
.gcal-chip:hover { filter: brightness(1.15); }
[data-theme="dark"] .gcal-chip { background: #1d4ed8; color: #bfdbfe; }
.gcal-chip-label {
    overflow: hidden; text-overflow: ellipsis; flex: 1;
}
.gcal-chip-n {
    flex-shrink: 0; font-size: 10px; font-weight: 700;
    background: rgba(255,255,255,.28); border-radius: 10px;
    padding: 0 4px; min-width: 17px; text-align: center;
}
[data-theme="dark"] .gcal-chip-n { background: rgba(255,255,255,.2); }
.gcal-mais {
    background: transparent; color: var(--text-2);
    border: 1px solid var(--border); font-style: italic;
    justify-content: center;
}
.gcal-mais:hover { filter: none; background: var(--bg-hover); }
"""

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Calendário — SSA Monitor</title>
{_JS_THEME_INIT}
<style>{CSS_BASE}
{css_cal}
</style>
</head>
<body>
    {_topbar("/calendario", usuario)}
    <div class="page-content">
        <div class="page-header">
            <h1>Calendário de Movimentações</h1>
            <p>Clique em qualquer processo para ver os detalhes</p>
        </div>

        <div class="nav-mes">
            <a class="nav-btn" href="/calendario?ano={ano_ant}&mes={mes_ant_str}">&larr; {MESES_PT[mes_ant]}</a>
            <span class="mes-atual">{MESES_PT[mes]} de {ano}</span>
            {nav_prox}
        </div>

        <div class="card" style="margin-bottom:16px;padding:16px 20px;display:flex;gap:32px;flex-wrap:wrap;">
            <div><span style="font-size:24px;font-weight:800;color:var(--blue)">{total_mes}</span>
                 <span style="color:var(--text-2);margin-left:6px;">processos com movimentação no mês</span></div>
            <div><span style="font-size:24px;font-weight:800;color:var(--green)">{dias_com_mov}</span>
                 <span style="color:var(--text-2);margin-left:6px;">dias com atividade</span></div>
        </div>

        <div class="gcal-wrap">
            <div class="gcal">
                {header_html}
                {cells_html}
            </div>
        </div>
    </div>
</body>
</html>"""


# ─────────────────────────────────────────────
# PÁGINA: /processo/<id>  — detalhe completo
# ─────────────────────────────────────────────
def gerar_html_detalhe_processo(processo_id: int, usuario=None):
    processo  = buscar_processo_por_id_dashboard(processo_id)
    if not processo:
        return None

    movimentacoes = buscar_movimentacoes_do_processo(processo_id)
    historico     = buscar_historico_consultas_do_processo(processo_id)

    num    = escape(str(processo.get("numero_processo") or ""))
    emp    = escape(str(processo.get("empresa") or "—"))
    orgao  = escape(str(processo.get("orgao") or ""))
    mun    = escape(str(processo.get("municipio") or orgao))
    status = _badge_status(str(processo.get("status_atual") or ""))
    robo   = escape(str(processo.get("robo") or "—"))
    dt_mov = escape(_fmt_data(processo.get("data_ultimo_movimento")))
    ult_c  = escape(_fmt_data(processo.get("ultima_consulta"), hora=True))
    url_o  = escape(str(processo.get("url_orgao") or ""))
    objeto = escape(str(processo.get("objeto") or ""))

    # Movimentações
    linhas_mov = ""
    for m in movimentacoes:
        dt   = escape(_fmt_data(m.get("data_movimento")))
        dc   = escape(_fmt_data(m.get("data_captura")))
        hr   = escape(str(m.get("hora_captura") or "")[:5])
        desc = escape(str(m.get("descricao") or "").strip())
        if not desc or len(desc) < 5:
            continue
        linhas_mov += f"""
        <tr>
            <td style="white-space:nowrap;color:#6b7280;font-size:13px;">{dt}</td>
            <td style="white-space:nowrap;color:#9ca3af;font-size:12px;">{dc} {hr}</td>
            <td style="white-space:normal;font-size:14px;">{desc}</td>
        </tr>"""
    if not linhas_mov:
        linhas_mov = '<tr><td colspan="3" class="vazio">Nenhuma movimentação registrada.</td></tr>'

    # Histórico de consultas
    linhas_hist = ""
    for h in historico:
        sc  = str(h.get("status_consulta") or "")
        dc  = escape(str(h.get("data_consulta") or ""))
        msg = escape(str(h.get("mensagem") or "")[:120])
        if sc == "OK":
            badge = f'<span class="badge verde">{escape(sc)}</span>'
        elif "NAO_ENCONTRADO" in sc or "ERRO" in sc:
            badge = f'<span class="badge laranja">{escape(sc)}</span>'
        else:
            badge = f'<span class="badge cinza">{escape(sc)}</span>'
        linhas_hist += f"""
        <tr>
            <td>{badge}</td>
            <td style="color:#6b7280;font-size:13px;">{dc}</td>
            <td style="white-space:normal;font-size:13px;color:#374151;">{msg}</td>
        </tr>"""
    if not linhas_hist:
        linhas_hist = '<tr><td colspan="3" class="vazio">Nenhuma consulta registrada.</td></tr>'

    link_portal = (
        f'<a href="{url_o}" target="_blank" style="color:#2563eb;font-size:13px;">'
        f'Abrir no portal da prefeitura ›</a>'
        if url_o else ""
    )

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Processo {num} · SSA Monitor</title>
    {_JS_THEME_INIT}
    <style>{CSS_BASE}</style>
</head>
<body>
    {_topbar("/movimentacoes-hoje", usuario)}
    <div class="page-content">
        <div class="page-header">
            <h1>Processo {num}</h1>
            <div class="subtitulo">{emp} · {orgao}</div>
        </div>

        <a href="javascript:history.back()" class="btn secundario" style="margin-bottom:20px;">
            &larr; Voltar
        </a>

        <div class="cards" style="grid-template-columns:repeat(3,minmax(160px,1fr));margin-bottom:22px;">
            <div class="card">
                <h2 style="font-size:18px;letter-spacing:-.5px;">{num}</h2>
                <p>Número do processo</p>
            </div>
            <div class="card">
                <h2 style="font-size:18px;letter-spacing:-.5px;">{mun}</h2>
                <p>Prefeitura / Município</p>
            </div>
            <div class="card sucesso">
                <h2 style="font-size:16px;letter-spacing:0;">{status}</h2>
                <p>Status atual</p>
            </div>
        </div>

        {"<section style='margin-bottom:20px;border-left:4px solid var(--blue);padding-left:20px;'><h2 style='margin-bottom:10px;'>Objeto</h2><p style='line-height:1.6;color:var(--text-1);white-space:pre-wrap;max-width:820px;'>" + objeto + "</p></section>" if objeto else ""}

        <section style="margin-bottom:20px;">
            <h2>Informações do processo</h2>
            <table style="width:auto;">
                <tr>
                    <td style="color:var(--text-2);padding:6px 16px 6px 0;white-space:nowrap;">Empresa</td>
                    <td><strong>{emp}</strong></td>
                </tr>
                <tr>
                    <td style="color:var(--text-2);padding:6px 16px 6px 0;white-space:nowrap;">Robô</td>
                    <td>{robo}</td>
                </tr>
                <tr>
                    <td style="color:var(--text-2);padding:6px 16px 6px 0;white-space:nowrap;">Último mov.</td>
                    <td>{dt_mov}</td>
                </tr>
                <tr>
                    <td style="color:var(--text-2);padding:6px 16px 6px 0;white-space:nowrap;">Última consulta</td>
                    <td>{ult_c}</td>
                </tr>
                {"<tr><td style='color:var(--text-2);padding:6px 16px 6px 0;white-space:nowrap;'>Portal</td><td>" + link_portal + "</td></tr>" if link_portal else ""}
            </table>
        </section>

        <section>
            <h2>Histórico de movimentações
                <span style="font-size:13px;color:var(--text-2);font-weight:400;">({len(movimentacoes)} registros)</span>
            </h2>
            <table>
                <thead>
                    <tr>
                        <th style="white-space:nowrap;">Data mov.</th>
                        <th style="white-space:nowrap;">Capturado em</th>
                        <th>Descrição</th>
                    </tr>
                </thead>
                <tbody>{linhas_mov}</tbody>
            </table>
        </section>

        <section>
            <h2>Histórico de consultas do robô
                <span style="font-size:13px;color:var(--text-2);font-weight:400;">(últimas {len(historico)})</span>
            </h2>
            <table>
                <thead>
                    <tr>
                        <th>Resultado</th>
                        <th>Data/hora</th>
                        <th>Mensagem</th>
                    </tr>
                </thead>
                <tbody>{linhas_hist}</tbody>
            </table>
        </section>

        <div class="footer">SSA Monitor Processos · Detalhe do Processo</div>
    </div>
    {_JS_THEME_TOGGLE}
</body>
</html>"""


# ─────────────────────────────────────────────
# PÁGINA: /processos
# ─────────────────────────────────────────────
def gerar_html_processos(processos, orgaos, empresas, statuses,
                         filtro_orgao="", filtro_empresa="", filtro_status="",
                         usuario=None):
    _STATUS_CORES = {
        "Em andamento": "#2563eb", "Em analise": "#f59e0b", "Em análise": "#f59e0b",
        "Indeferido": "#ef4444", "Deferido": "#10b981", "Finalizado": "#10b981",
        "Encerrado": "#6b7280",
    }

    def _badge_status(s):
        if not s:
            return "<span style='color:var(--text-3);font-size:12px;'>—</span>"
        cor = _STATUS_CORES.get(s, "#6b7280")
        return f"<span style='background:{cor}22;color:{cor};padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600;'>{escape(s)}</span>"

    def _badge_resultado(r):
        if not r:
            return "<span style='color:var(--text-3);font-size:12px;'>—</span>"
        if r == "OK":
            return "<span style='background:#10b98122;color:#10b981;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600;'>OK</span>"
        return "<span style='background:#ef444422;color:#ef4444;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600;'>ERRO_CONSULTA</span>"

    def _opt(lista, sel):
        return "".join(f"<option value='{escape(v)}' {'selected' if v==sel else ''}>{escape(v)}</option>" for v in lista)

    linhas = ""
    for p in processos:
        pid    = p["id"]
        num    = escape(str(p["numero_processo"] or ""))
        emp    = escape(str(p["empresa"] or ""))
        cli    = escape(str(p["cliente"] or ""))
        org    = escape(str(p["orgao"] or ""))
        ult_c  = _fmt_data(p["ultima_consulta"], hora=True) if p.get("ultima_consulta") else "—"
        ult_m  = _fmt_data(p["data_ultimo_movimento"]) if p.get("data_ultimo_movimento") else "—"
        badge_s = _badge_status(p.get("status_atual"))
        badge_r = _badge_resultado(p.get("ultimo_resultado"))
        linhas += f"""
        <tr onclick="window.location='/processo/{pid}'" style="cursor:pointer;">
            <td><strong style='font-size:12px;'>{num}</strong></td>
            <td><span style='font-size:12px;'>{emp}</span><span style='font-size:10px;color:var(--text-3);margin-left:5px;'>({cli})</span></td>
            <td style='font-size:12px;'>{org}</td>
            <td>{badge_s}</td>
            <td>{badge_r}</td>
            <td style='color:var(--text-2);'>{ult_c}</td>
            <td style='color:var(--text-2);'>{ult_m}</td>
        </tr>"""

    topbar = _topbar("/processos", usuario)
    total  = len(processos)

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Processos — SSA Monitor</title>
{_JS_THEME_INIT}
<style>{CSS_BASE}
table {{ width:100%;border-collapse:collapse; }}
th {{ text-align:left;font-size:10px;font-weight:600;color:var(--text-3);
     text-transform:uppercase;letter-spacing:.05em;padding:7px 10px;
     border-bottom:1px solid var(--border);white-space:nowrap; }}
td {{ padding:7px 10px;border-bottom:1px solid var(--border);vertical-align:middle;font-size:12px; }}
tr:last-child td {{ border-bottom:none; }}
tr:hover td {{ background:var(--bg-hover); }}
select {{ background:var(--bg-card);color:var(--text-1);border:1px solid var(--border);
          border-radius:7px;padding:5px 9px;font-size:12px;font-family:inherit;cursor:pointer; }}
</style>
</head>
<body>
{topbar}
<div class="container" style="max-width:1200px;">
    <div style="display:flex;align-items:center;justify-content:space-between;
                flex-wrap:wrap;gap:10px;margin-bottom:16px;">
        <div>
            <h1 style="font-size:22px;font-weight:800;">Processos</h1>
            <p style="color:var(--text-2);font-size:12px;margin-top:2px;">
                {total} processo{'s' if total != 1 else ''} ativo{'s' if total != 1 else ''}
            </p>
        </div>
        <form method="get" action="/processos"
              style="display:flex;gap:6px;flex-wrap:wrap;align-items:center;">
            <select name="orgao" onchange="this.form.submit()">
                <option value="">Todas as prefeituras</option>
                {_opt(orgaos, filtro_orgao)}
            </select>
            <select name="empresa" onchange="this.form.submit()">
                <option value="">Todas as empresas</option>
                {_opt(empresas, filtro_empresa)}
            </select>
            <select name="status" onchange="this.form.submit()">
                <option value="">Todos os status</option>
                {_opt(statuses, filtro_status)}
            </select>
            {'<a href="/processos" style="font-size:11px;color:var(--text-3);">Limpar</a>' if (filtro_orgao or filtro_empresa or filtro_status) else ''}
        </form>
    </div>

    <div style="background:var(--bg-card);border:1px solid var(--border);
                border-radius:12px;overflow:hidden;overflow-x:auto;">
        <table>
            <thead>
                <tr>
                    <th>Nº Processo</th>
                    <th>Empresa</th>
                    <th>Prefeitura</th>
                    <th>Status</th>
                    <th>Resultado</th>
                    <th>Última consulta</th>
                    <th>Último mov.</th>
                </tr>
            </thead>
            <tbody>
                {linhas if linhas else "<tr><td colspan='7' style='text-align:center;padding:32px;color:var(--text-3);'>Nenhum processo encontrado.</td></tr>"}
            </tbody>
        </table>
    </div>
</div>
</body>
</html>"""


# ─────────────────────────────────────────────
# PÁGINA: /relatorio  — filtros
# ─────────────────────────────────────────────
def gerar_html_relatorio(orgaos, empresas, statuses, usuario=None):
    hoje = date.today()
    data_fim_d = str(hoje)
    data_inicio_d = str(hoje.replace(day=1))

    def _opts(lista, placeholder):
        h = f"<option value=''>{placeholder}</option>"
        for item in lista:
            h += f"<option value='{escape(item)}'>{escape(item)}</option>"
        return h

    sel_style = ("padding:8px 12px;border:1px solid var(--border);border-radius:8px;"
                 "background:var(--bg-card);color:var(--text-1);font-size:13px;"
                 "cursor:pointer;min-width:170px;font-family:inherit;")
    inp_style = ("padding:8px 12px;border:1px solid var(--border);border-radius:8px;"
                 "background:var(--bg-card);color:var(--text-1);font-size:13px;font-family:inherit;")
    lbl_style = "display:block;font-size:12px;font-weight:600;color:var(--text-2);margin-bottom:5px;"

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Relatório · SSA Monitor</title>
    {_JS_THEME_INIT}
    <style>{CSS_BASE}</style>
</head>
<body>
    {_topbar("/relatorio", usuario)}
    <div class="page-content">
        <div class="page-header">
            <h1>Relatório de Processos</h1>
            <div class="subtitulo">Filtre por período, prefeitura, empresa e status · visualize ou exporte em Excel</div>
        </div>

        <section>
            <h2>Filtros</h2>
            <form id="frm" style="display:flex;flex-wrap:wrap;gap:18px;align-items:flex-end;">
                <div>
                    <label style="{lbl_style}">Data inicial</label>
                    <input type="date" name="data_inicio" value="{data_inicio_d}" max="{data_fim_d}" style="{inp_style}">
                </div>
                <div>
                    <label style="{lbl_style}">Data final</label>
                    <input type="date" name="data_fim" value="{data_fim_d}" max="{data_fim_d}" style="{inp_style}">
                </div>
                <div>
                    <label style="{lbl_style}">Prefeitura</label>
                    <select name="orgao" style="{sel_style}">{_opts(orgaos, 'Todas as prefeituras')}</select>
                </div>
                <div>
                    <label style="{lbl_style}">Empresa / Cliente</label>
                    <select name="empresa" style="{sel_style}">{_opts(empresas, 'Todas as empresas')}</select>
                </div>
                <div>
                    <label style="{lbl_style}">Status</label>
                    <select name="status" style="{sel_style}">{_opts(statuses, 'Todos os status')}</select>
                </div>
                <div style="display:flex;gap:10px;flex-wrap:wrap;">
                    <button type="button" onclick="ir('/relatorio/visualizar')" class="btn">
                        &#128438; Visualizar Relatório
                    </button>
                    <button type="button" onclick="ir('/relatorio/exportar-excel')" class="btn secundario">
                        &#128229; Exportar Excel
                    </button>
                </div>
            </form>
        </section>

        <section style="background:var(--bg-th);border-style:dashed;">
            <h2 style="margin-bottom:10px;">O que cada opção gera</h2>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;">
                <div>
                    <strong style="font-size:13px;">&#128438; Visualizar Relatório</strong>
                    <p style="color:var(--text-2);font-size:13px;margin-top:4px;">
                        Abre uma página formatada, agrupada por prefeitura, com as movimentações
                        detectadas no período. Pode ser impressa ou salva como PDF pelo navegador.
                    </p>
                </div>
                <div>
                    <strong style="font-size:13px;">&#128229; Exportar Excel</strong>
                    <p style="color:var(--text-2);font-size:13px;margin-top:4px;">
                        Baixa um arquivo <code>.xlsx</code> com 3 abas: <em>Resumo</em> (totais por prefeitura),
                        <em>Movimentações</em> (lista detalhada) e <em>Processos</em> (visão geral de todos).
                    </p>
                </div>
            </div>
        </section>
    </div>
    <script>
    function ir(base) {{
        const p = new URLSearchParams(new FormData(document.getElementById('frm')));
        window.location = base + '?' + p.toString();
    }}
    </script>
    {_JS_THEME_TOGGLE}
</body>
</html>"""


# ─────────────────────────────────────────────
# PÁGINA: /relatorio/visualizar  — relatório imprimível
# ─────────────────────────────────────────────
def gerar_html_relatorio_visualizar(por_orgao, processos, filtros, usuario=None):
    data_inicio = filtros.get('data_inicio', '')
    data_fim    = filtros.get('data_fim', '')
    orgao_f     = filtros.get('orgao', '')
    empresa_f   = filtros.get('empresa', '')
    status_f    = filtros.get('status', '')

    total_processos = len(processos)
    com_mov   = sum(1 for p in processos if p['movimentacoes'])
    sem_mov   = total_processos - com_mov
    total_mov = sum(len(p['movimentacoes']) for p in processos)

    filtros_desc = [v for v in [
        f"Prefeitura: {orgao_f}" if orgao_f else "",
        f"Empresa: {empresa_f}" if empresa_f else "",
        f"Status: {status_f}" if status_f else "",
    ] if v]
    filtros_txt = " · ".join(filtros_desc) if filtros_desc else ""

    qs_parts = "&".join(f"{k}={escape(v)}" for k, v in filtros.items() if v)

    secoes = ""
    for orgao_nome, procs in sorted(por_orgao.items()):
        com_mov_org = sum(1 for p in procs if p['movimentacoes'])

        # Processos COM movimentação
        linhas_com = ""
        for p in [x for x in procs if x['movimentacoes']]:
            num  = escape(str(p.get('numero_processo') or ''))
            emp  = escape(str(p.get('empresa') or ''))
            st   = str(p.get('status_atual') or '')
            n_m  = len(p['movimentacoes'])
            obj  = escape(str(p.get('objeto') or ''))

            movs_html = ""
            for m in p['movimentacoes'][:15]:
                dt   = escape(_fmt_data(m.get('data_movimento')))
                desc = escape(str(m.get('descricao') or '').strip()[:300])
                if not desc or len(desc) < 5:
                    continue
                movs_html += f"""
                <tr style="background:var(--bg-th);">
                    <td style="font-size:12px;color:var(--text-2);padding:4px 10px;white-space:nowrap;">{dt}</td>
                    <td style="font-size:12px;padding:4px 10px;white-space:normal;max-width:480px;">{desc}</td>
                </tr>"""

            obj_row = (f"<tr><td colspan='4' style='padding:2px 12px 6px;font-size:12px;"
                       f"color:var(--text-2);white-space:normal;max-width:600px;'>"
                       f"<em>{obj}</em></td></tr>") if obj else ""

            linhas_com += f"""
            <tr style="background:#f0fdf4;">
                <td style="font-weight:700;padding:8px 12px;border-left:3px solid var(--green);">{num}</td>
                <td style="padding:8px 12px;">{emp}</td>
                <td style="padding:8px 12px;">{_badge_status(st)}</td>
                <td style="padding:8px 12px;text-align:center;font-weight:700;color:var(--green);">{n_m}</td>
            </tr>
            {obj_row}
            <tr>
                <td colspan="4" style="padding:0 12px 10px 28px;">
                    <table style="width:100%;border-collapse:collapse;">
                        <thead><tr>
                            <th style="font-size:11px;color:var(--text-3);text-align:left;padding:3px 10px;white-space:nowrap;">Data mov.</th>
                            <th style="font-size:11px;color:var(--text-3);text-align:left;padding:3px 10px;">Descrição</th>
                        </tr></thead>
                        <tbody>{movs_html if movs_html else '<tr><td colspan="2" style="color:var(--text-3);padding:4px 10px;font-size:12px;">Sem descrição disponível.</td></tr>'}</tbody>
                    </table>
                </td>
            </tr>"""

        # Processos SEM movimentação
        linhas_sem = ""
        for p in [x for x in procs if not x['movimentacoes']]:
            num    = escape(str(p.get('numero_processo') or ''))
            emp    = escape(str(p.get('empresa') or ''))
            st     = str(p.get('status_atual') or '')
            dt_mov = escape(_fmt_data(p.get('data_ultimo_movimento')))
            linhas_sem += f"""
            <tr style="opacity:.55;">
                <td style="padding:5px 12px;">{num}</td>
                <td style="padding:5px 12px;">{emp}</td>
                <td style="padding:5px 12px;">{_badge_status(st)}</td>
                <td style="padding:5px 12px;color:var(--text-3);">{dt_mov}</td>
            </tr>"""

        bloco_com = ""
        if linhas_com:
            bloco_com = f"""
            <p style="font-size:12px;font-weight:600;color:var(--green);margin:14px 0 6px;">
                &#10004; Com movimentação no período ({com_mov_org})
            </p>
            <table style="width:100%;border-collapse:collapse;margin-bottom:10px;">
                <thead><tr>
                    <th style="text-align:left;padding:7px 12px;background:var(--bg-th);font-size:11px;color:var(--text-2);text-transform:uppercase;letter-spacing:.05em;">Processo</th>
                    <th style="text-align:left;padding:7px 12px;background:var(--bg-th);font-size:11px;color:var(--text-2);text-transform:uppercase;letter-spacing:.05em;">Empresa</th>
                    <th style="text-align:left;padding:7px 12px;background:var(--bg-th);font-size:11px;color:var(--text-2);text-transform:uppercase;letter-spacing:.05em;">Status</th>
                    <th style="text-align:center;padding:7px 12px;background:var(--bg-th);font-size:11px;color:var(--text-2);text-transform:uppercase;letter-spacing:.05em;">Movs.</th>
                </tr></thead>
                <tbody>{linhas_com}</tbody>
            </table>"""

        bloco_sem = ""
        sem_count = len(procs) - com_mov_org
        if linhas_sem:
            bloco_sem = f"""
            <details style="margin-top:6px;" class="sem-mov">
                <summary style="font-size:12px;color:var(--text-3);cursor:pointer;font-weight:500;user-select:none;">
                    &#8212; Sem movimentação no período ({sem_count}) — clique para expandir
                </summary>
                <table style="width:100%;border-collapse:collapse;margin-top:8px;">
                    <thead><tr>
                        <th style="text-align:left;padding:6px 12px;background:var(--bg-th);font-size:11px;color:var(--text-3);">Processo</th>
                        <th style="text-align:left;padding:6px 12px;background:var(--bg-th);font-size:11px;color:var(--text-3);">Empresa</th>
                        <th style="text-align:left;padding:6px 12px;background:var(--bg-th);font-size:11px;color:var(--text-3);">Status</th>
                        <th style="text-align:left;padding:6px 12px;background:var(--bg-th);font-size:11px;color:var(--text-3);">Último mov.</th>
                    </tr></thead>
                    <tbody>{linhas_sem}</tbody>
                </table>
            </details>"""

        secoes += f"""
        <section style="page-break-inside:avoid;">
            <h2 style="font-size:15px;font-weight:700;border-bottom:2px solid var(--border);padding-bottom:8px;margin-bottom:4px;">
                {escape(orgao_nome)}
                <span style="font-size:12px;font-weight:400;color:var(--text-2);margin-left:8px;">
                    {len(procs)} processo{'s' if len(procs) != 1 else ''} &middot; {com_mov_org} com movimentação
                </span>
            </h2>
            {bloco_com}
            {bloco_sem}
        </section>"""

    if not secoes:
        secoes = '<section><p class="vazio">Nenhum processo encontrado para os filtros selecionados.</p></section>'

    di_fmt  = _fmt_data(data_inicio) if data_inicio else "—"
    df_fmt  = _fmt_data(data_fim)    if data_fim    else "—"
    agora   = datetime.now().strftime("%d/%m/%Y - %H:%M")

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Relatório · SSA Monitor</title>
    {_JS_THEME_INIT}
    <style>{CSS_BASE}
    @media print {{
        .topbar, .no-print {{ display: none !important; }}
        body {{ background: white !important; }}
        section {{ page-break-inside: avoid; border: none !important; box-shadow: none !important; }}
        .page-content {{ max-width: 100% !important; padding: 0 !important; }}
        details.sem-mov {{ display: block; }}
        details.sem-mov > summary {{ display: none; }}
        details.sem-mov > table {{ display: table !important; }}
    }}
    </style>
</head>
<body>
    {_topbar("/relatorio", usuario)}
    <div class="page-content">

        <div class="no-print" style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:10px;">
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a href="/relatorio" class="btn secundario">&larr; Voltar aos filtros</a>
                <a href="/relatorio/exportar-excel?{qs_parts}" class="btn secundario">&#128229; Exportar Excel</a>
            </div>
            <button onclick="window.print()" class="btn">&#128438; Imprimir / Salvar PDF</button>
        </div>

        <!-- CABEÇALHO IMPRIMÍVEL -->
        <div style="text-align:center;margin-bottom:24px;padding:22px 24px;
                    background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius);">
            <div style="font-size:10px;font-weight:700;color:var(--text-3);letter-spacing:.12em;
                        text-transform:uppercase;margin-bottom:6px;">SSA Monitor Processos</div>
            <h1 style="font-size:20px;font-weight:800;margin-bottom:6px;letter-spacing:-.3px;">
                Relatório de Monitoramento de Processos
            </h1>
            <div style="color:var(--text-2);font-size:14px;">
                Período: <strong>{di_fmt}</strong> a <strong>{df_fmt}</strong>
            </div>
            {f'<div style="color:var(--text-2);font-size:13px;margin-top:3px;">{escape(filtros_txt)}</div>' if filtros_txt else ''}
            <div style="color:var(--text-3);font-size:12px;margin-top:8px;">Gerado em {agora}</div>
        </div>

        <!-- CARDS DE RESUMO -->
        <div class="cards" style="grid-template-columns:repeat(4,minmax(130px,1fr));margin-bottom:24px;">
            <div class="card"><h2>{total_processos}</h2><p>Processos analisados</p></div>
            <div class="card sucesso"><h2>{com_mov}</h2><p>Com movimentação</p></div>
            <div class="card neutro"><h2>{sem_mov}</h2><p>Sem movimentação</p></div>
            <div class="card alerta"><h2>{total_mov}</h2><p>Total de movimentações</p></div>
        </div>

        <!-- SEÇÕES POR PREFEITURA -->
        {secoes}

        <div class="footer no-print">SSA Monitor Processos · Relatório</div>
    </div>
    {_JS_THEME_TOGGLE}
</body>
</html>"""


# ─────────────────────────────────────────────
# GERADOR: Excel (.xlsx)
# ─────────────────────────────────────────────
def _gerar_excel_relatorio(por_orgao, processos, filtros):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = Workbook()

    AZUL    = "0F172A"
    BORDA_C = Side(style='thin', color='E2E8F0')
    brd     = Border(left=BORDA_C, right=BORDA_C, top=BORDA_C, bottom=BORDA_C)

    def hdr(cell, bg=AZUL):
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = brd

    def dado(cell, bold=False, cor=None):
        cell.font      = Font(bold=bold, color=cor or "0F172A", size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border    = brd

    data_inicio = filtros.get('data_inicio', '')
    data_fim    = filtros.get('data_fim', '')
    total       = len(processos)
    com_mov     = sum(1 for p in processos if p['movimentacoes'])
    total_movs  = sum(len(p['movimentacoes']) for p in processos)

    # ── Aba 1: Resumo ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"

    ws1.merge_cells('A1:C1')
    ws1['A1'] = "SSA Monitor Processos — Relatório de Monitoramento"
    ws1['A1'].font      = Font(bold=True, size=13, color=AZUL)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.row_dimensions[1].height = 24

    for r, (k, v) in enumerate([
        ("Período", f"{_fmt_data(data_inicio)} a {_fmt_data(data_fim)}"),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y - %H:%M")),
        ("Prefeitura", filtros.get('orgao') or "Todas"),
        ("Empresa", filtros.get('empresa') or "Todas"),
        ("Status", filtros.get('status') or "Todos"),
    ], start=2):
        ws1.cell(r, 1, k).font = Font(bold=True, size=10)
        ws1.cell(r, 2, v).font = Font(size=10)

    ws1.append([])

    # Totais gerais
    for col, h in enumerate(["Indicador", "Valor"], 1):
        hdr(ws1.cell(ws1.max_row + 1, col, h))
    for label, val in [
        ("Total de processos analisados", total),
        ("Com movimentação no período",   com_mov),
        ("Sem movimentação no período",   total - com_mov),
        ("Total de movimentações detectadas", total_movs),
    ]:
        ws1.append([label, val])
        row = ws1.max_row
        dado(ws1.cell(row, 1))
        dado(ws1.cell(row, 2), bold=True)

    ws1.append([])

    # Por prefeitura
    for col, h in enumerate(["Prefeitura", "Processos", "Com mov.", "Sem mov.", "Movimentações"], 1):
        hdr(ws1.cell(ws1.max_row + 1, col, h))
    for orgao_nome, procs in sorted(por_orgao.items()):
        cm = sum(1 for p in procs if p['movimentacoes'])
        tm = sum(len(p['movimentacoes']) for p in procs)
        ws1.append([orgao_nome, len(procs), cm, len(procs) - cm, tm])
        row = ws1.max_row
        for col in range(1, 6):
            dado(ws1.cell(row, col), bold=(col == 3 and cm > 0), cor=("166534" if (col == 3 and cm > 0) else None))

    for col, w in zip(range(1, 6), [32, 12, 12, 12, 16]):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = 'A2'

    # ── Aba 2: Movimentações ───────────────────────────────────────
    ws2 = wb.create_sheet("Movimentações")
    h2 = ["Prefeitura", "Empresa", "Nº Processo", "Status", "Data Mov.", "Descrição"]
    for col, h in enumerate(h2, 1):
        hdr(ws2.cell(1, col, h))

    for p in processos:
        for m in p['movimentacoes']:
            ws2.append([
                p.get('orgao', ''),
                p.get('empresa', ''),
                p.get('numero_processo', ''),
                p.get('status_atual', ''),
                str(m.get('data_movimento', '')),
                str(m.get('descricao', '') or '').strip()[:500],
            ])
            row = ws2.max_row
            for col in range(1, 7):
                dado(ws2.cell(row, col))

    if ws2.max_row == 1:
        ws2.append(["Nenhuma movimentação no período selecionado."])

    for col, w in zip(range(1, 7), [22, 20, 14, 18, 13, 70]):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A2'

    # ── Aba 3: Processos ───────────────────────────────────────────
    ws3 = wb.create_sheet("Processos")
    h3 = ["Prefeitura", "Empresa", "Nº Processo", "Status", "Movs. no período", "Último mov.", "Objeto"]
    for col, h in enumerate(h3, 1):
        hdr(ws3.cell(1, col, h))

    for p in processos:
        n_m = len(p['movimentacoes'])
        ws3.append([
            p.get('orgao', ''),
            p.get('empresa', ''),
            p.get('numero_processo', ''),
            p.get('status_atual', ''),
            n_m,
            str(p.get('data_ultimo_movimento', '') or ''),
            str(p.get('objeto', '') or '').strip()[:300],
        ])
        row = ws3.max_row
        for col in range(1, 8):
            dado(ws3.cell(row, col), bold=(col == 5 and n_m > 0), cor=("166534" if (col == 5 and n_m > 0) else None))

    for col, w in zip(range(1, 8), [22, 20, 14, 18, 16, 14, 50]):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# CSS da tela de autenticação (fora do layout com topbar)
# ─────────────────────────────────────────────
CSS_AUTH = """
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    body {
        font-family: 'Inter', Arial, sans-serif;
        background: #0f172a;
        color: #0f172a;
        min-height: 100vh;
        display: flex; align-items: center; justify-content: center;
        font-size: 14px;
    }
    .auth-box {
        background: #fff;
        border-radius: 12px;
        padding: 32px 34px;
        width: 100%; max-width: 360px;
        box-shadow: 0 8px 28px rgba(0,0,0,.25);
    }
    .auth-box h1 { font-size: 18px; font-weight: 700; margin-bottom: 4px; }
    .auth-box .sub { color: #64748b; font-size: 13px; margin-bottom: 20px; }
    .auth-box label { display: block; font-size: 12px; font-weight: 600; color: #334155; margin-bottom: 5px; margin-top: 14px; }
    .auth-box input {
        width: 100%; padding: 9px 12px;
        border: 1px solid #e2e8f0; border-radius: 8px;
        font-size: 14px; font-family: inherit;
    }
    .auth-box button {
        width: 100%; margin-top: 20px;
        padding: 10px; border: none; border-radius: 8px;
        background: #2563eb; color: #fff;
        font-size: 14px; font-weight: 600; cursor: pointer;
    }
    .auth-box button:hover { opacity: .9; }
    .auth-erro {
        background: #fef2f2; color: #b91c1c;
        border: 1px solid #fecaca; border-radius: 8px;
        padding: 9px 12px; font-size: 13px; margin-top: 14px;
    }
"""


def gerar_html_login(erro=None):
    erro_html = f'<div class="auth-erro">{escape(erro)}</div>' if erro else ""

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Entrar · SSA Monitor</title>
    <style>{CSS_AUTH}</style>
</head>
<body>
    <div class="auth-box">
        <h1>SSA Monitor Processos</h1>
        <div class="sub">Acesso restrito — uso exclusivo do escritório</div>
        <form method="POST" action="/login">
            <label>E-mail</label>
            <input type="email" name="email" required autofocus>
            <label>Senha</label>
            <input type="password" name="senha" required>
            <button type="submit">Entrar</button>
        </form>
        {erro_html}
    </div>
</body>
</html>"""


def gerar_html_trocar_senha(token_csrf, erro=None, obrigatorio=False):
    erro_html = f'<div class="auth-erro">{escape(erro)}</div>' if erro else ""
    aviso = (
        '<div class="sub">Defina uma senha nova antes de continuar.</div>'
        if obrigatorio else
        '<div class="sub">Trocar minha senha</div>'
    )

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Trocar senha · SSA Monitor</title>
    <style>{CSS_AUTH}</style>
</head>
<body>
    <div class="auth-box">
        <h1>Trocar senha</h1>
        {aviso}
        <form method="POST" action="/trocar-senha">
            <input type="hidden" name="csrf" value="{escape(token_csrf)}">
            <label>Senha atual</label>
            <input type="password" name="senha_atual" required>
            <label>Nova senha</label>
            <input type="password" name="senha_nova" required minlength="8">
            <label>Confirmar nova senha</label>
            <input type="password" name="senha_confirma" required minlength="8">
            <button type="submit">Salvar</button>
        </form>
        {erro_html}
    </div>
</body>
</html>"""


def gerar_html_admin_usuarios(usuario, usuarios, token_csrf, mensagem=None):
    mensagem_html = (
        f'<div style="background:#dbeafe;color:#1e40af;border-radius:8px;padding:10px 14px;font-size:13px;margin-bottom:16px;">{escape(mensagem)}</div>'
        if mensagem else ""
    )

    eu_id = usuario.get("usuario_id")
    linhas = ""
    for u in usuarios:
        status = "Ativo" if u.get("ativo") else "Desativado"
        badge_cls = "verde" if u.get("ativo") else "cinza"
        papel = "Administrador" if u.get("is_admin") else "Usuário"
        acao_status = "desativar" if u.get("ativo") else "ativar"
        ultimo_login = escape(str(u.get("ultimo_login") or "—"))
        proprio = u.get("id") == eu_id

        btn_excluir = ""
        if not proprio:
            nome_u = escape(u.get("nome") or u.get("email") or "este usuário")
            btn_excluir = f"""
                <form method="POST" action="/admin/usuarios/{u.get('id')}/excluir" style="display:inline;"
                      onsubmit="return confirm('Excluir {nome_u}? Esta ação é permanente e não pode ser desfeita.');">
                    <input type="hidden" name="csrf" value="{escape(token_csrf)}">
                    <button type="submit" class="btn secundario"
                            style="padding:4px 10px;font-size:12px;margin:0;color:#dc2626;border-color:#fca5a5;">
                        Excluir
                    </button>
                </form>"""

        linhas += f"""
        <tr>
            <td>{escape(u.get('nome') or '')}</td>
            <td>{escape(u.get('email') or '')}</td>
            <td>{escape(papel)}</td>
            <td><span class="badge {badge_cls}">{status}</span></td>
            <td>{ultimo_login}</td>
            <td style="white-space:nowrap;">
                <form method="POST" action="/admin/usuarios/{u.get('id')}/status" style="display:inline;">
                    <input type="hidden" name="csrf" value="{escape(token_csrf)}">
                    <button type="submit" class="btn secundario" style="padding:4px 10px;font-size:12px;margin:0;">{acao_status.capitalize()}</button>
                </form>
                <form method="POST" action="/admin/usuarios/{u.get('id')}/resetar-senha" style="display:inline;">
                    <input type="hidden" name="csrf" value="{escape(token_csrf)}">
                    <button type="submit" class="btn secundario" style="padding:4px 10px;font-size:12px;margin:0;">Resetar senha</button>
                </form>
                {btn_excluir}
            </td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Usuários · SSA Monitor</title>
    {_JS_THEME_INIT}
    <style>{CSS_BASE}</style>
</head>
<body>
    {_topbar("/admin/usuarios", usuario)}
    <div class="page-content">
        <div class="page-header">
            <h1>Usuários</h1>
            <div class="subtitulo">Gerencie quem tem acesso ao dashboard</div>
        </div>

        {mensagem_html}

        <section>
            <h2>Novo usuário</h2>
            <form method="POST" action="/admin/usuarios/criar" style="display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap;">
                <input type="hidden" name="csrf" value="{escape(token_csrf)}">
                <div>
                    <label style="display:block;font-size:12px;font-weight:600;color:var(--text-2);margin-bottom:4px;">Nome</label>
                    <input type="text" name="nome" required style="padding:8px 10px;border:1px solid var(--border);border-radius:8px;background:var(--bg-card);color:var(--text-1);">
                </div>
                <div>
                    <label style="display:block;font-size:12px;font-weight:600;color:var(--text-2);margin-bottom:4px;">E-mail</label>
                    <input type="email" name="email" required style="padding:8px 10px;border:1px solid var(--border);border-radius:8px;background:var(--bg-card);color:var(--text-1);">
                </div>
                <div>
                    <label style="display:block;font-size:12px;font-weight:600;color:var(--text-2);margin-bottom:4px;">Senha temporária</label>
                    <input type="text" name="senha_temporaria" required minlength="8" style="padding:8px 10px;border:1px solid var(--border);border-radius:8px;background:var(--bg-card);color:var(--text-1);">
                </div>
                <label style="display:flex;align-items:center;gap:6px;font-size:13px;color:var(--text-2);">
                    <input type="checkbox" name="is_admin" value="1" style="width:auto;"> Administrador
                </label>
                <button type="submit" class="btn">Criar usuário</button>
            </form>
        </section>

        <section>
            <h2>Usuários cadastrados</h2>
            <table>
                <thead>
                    <tr>
                        <th>Nome</th><th>E-mail</th><th>Papel</th><th>Status</th><th>Último login</th><th>Ações</th>
                    </tr>
                </thead>
                <tbody>{linhas}</tbody>
            </table>
        </section>
    </div>
    {_JS_THEME_TOGGLE}
</body>
</html>"""


# ─────────────────────────────────────────────
# HTTP Handler
# ─────────────────────────────────────────────
log = get_logger("dashboard")


class DashboardHandler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass

    # ── Helpers de infraestrutura ─────────────────────────────────
    def _ip_cliente(self):
        return self.headers.get("X-Real-IP") or self.client_address[0]

    def _sessao_atual(self):
        token = auth.extrair_cookie(self.headers.get("Cookie", ""), auth.NOME_COOKIE_SESSAO)
        if not token:
            return None, None
        sessao = auth_repository.buscar_sessao_valida(token)
        return token, sessao

    def _ler_corpo_form(self):
        tamanho = int(self.headers.get("Content-Length", 0) or 0)
        corpo = self.rfile.read(tamanho).decode("utf-8") if tamanho else ""
        bruto = parse_qs(corpo)
        return {chave: valores[0] for chave, valores in bruto.items()}

    def _responder_html(self, html, status=200, cookies=None):
        self.send_response(status)
        self.send_header("Content-type", "text/html; charset=utf-8")
        for cookie in (cookies or []):
            self.send_header("Set-Cookie", cookie)
        self.end_headers()
        self.wfile.write(html.encode("utf-8"))

    def _responder_json(self, dados: dict, status=200):
        corpo = json.dumps(dados, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-type", "application/json; charset=utf-8")
        self.end_headers()
        self.wfile.write(corpo)

    def _responder_arquivo(self, dados: bytes, content_type: str, nome_arquivo: str):
        self.send_response(200)
        self.send_header("Content-type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{nome_arquivo}"')
        self.send_header("Content-Length", str(len(dados)))
        self.end_headers()
        self.wfile.write(dados)

    def _redirecionar(self, local, cookies=None):
        self.send_response(302)
        self.send_header("Location", local)
        for cookie in (cookies or []):
            self.send_header("Set-Cookie", cookie)
        self.end_headers()

    def _negar(self, status=403, mensagem="Acesso negado."):
        self._responder_html(f"<h2>{escape(mensagem)}</h2>", status=status)

    # ── GET ────────────────────────────────────────────────────────
    def do_GET(self):
        parsed = urlparse(self.path)
        rota = parsed.path
        qs = parse_qs(parsed.query)
        token, sessao = self._sessao_atual()

        try:
            if rota == "/login":
                if sessao:
                    return self._redirecionar("/")
                return self._responder_html(gerar_html_login())

            if rota == "/logout":
                if token:
                    auth_repository.invalidar_sessao(token)
                return self._redirecionar("/login", cookies=[auth.montar_cookie_expirado()])

            if not sessao:
                return self._redirecionar("/login")

            if sessao.get("precisa_trocar_senha") and rota != "/trocar-senha":
                return self._redirecionar("/trocar-senha")

            if rota == "/trocar-senha":
                token_csrf = auth.gerar_token_csrf(token)
                return self._responder_html(
                    gerar_html_trocar_senha(token_csrf, obrigatorio=sessao.get("precisa_trocar_senha"))
                )

            if rota == "/admin/usuarios":
                if not sessao.get("is_admin"):
                    return self._negar()
                token_csrf = auth.gerar_token_csrf(token)
                usuarios = auth_repository.listar_usuarios()
                return self._responder_html(gerar_html_admin_usuarios(sessao, usuarios, token_csrf))

            if rota == "/api/admin/status-monitoramento":
                if not sessao.get("is_admin"):
                    return self._negar()
                status = _ler_status_monitoramento()
                ultima = buscar_ultima_execucao()
                ultima_fmt = _fmt_data(ultima, hora=True) if ultima else None
                return self._responder_json({
                    "running":       status.get("running", False),
                    "ultima_execucao": ultima_fmt,
                    "concluidos":    status.get("concluidos", 0),
                    "total":         status.get("total", 0),
                    "orgao_atual":   status.get("orgao_atual", ""),
                })

            if rota == "/":
                return self._responder_html(gerar_html_dashboard(sessao))

            if rota == "/calendario":
                hoje_cal = date.today()
                try:
                    ano_p = int(qs.get("ano", [hoje_cal.year])[0])
                    mes_p = int(qs.get("mes", [hoje_cal.month])[0])
                    if not (1 <= mes_p <= 12):
                        mes_p = hoje_cal.month
                except (ValueError, TypeError):
                    ano_p, mes_p = hoje_cal.year, hoje_cal.month
                return self._responder_html(gerar_html_calendario(ano_p, mes_p, sessao))

            if rota == "/movimentacoes-hoje":
                data_param = qs.get("data", [None])[0]
                return self._responder_html(gerar_html_movimentacoes_hoje(data_param, sessao))

            if rota.startswith("/processo/"):
                partes = rota.rstrip("/").split("/")
                pid_str = partes[-1] if partes else ""
                if not pid_str.isdigit():
                    return self._negar(400, "Requisição inválida.")
                html = gerar_html_detalhe_processo(int(pid_str), sessao)
                if html is None:
                    return self._negar(404, "Processo não encontrado.")
                return self._responder_html(html)

            if rota == "/processos":
                qs = parse_qs(urlparse(self.path).query)
                f_orgao   = qs.get("orgao",   [""])[0]
                f_empresa = qs.get("empresa", [""])[0]
                f_status  = qs.get("status",  [""])[0]
                orgaos, empresas, statuses = buscar_filtros_processos()
                processos = buscar_todos_processos(
                    f_orgao or None, f_empresa or None, f_status or None
                )
                return self._responder_html(gerar_html_processos(
                    processos, orgaos, empresas, statuses,
                    f_orgao, f_empresa, f_status, sessao
                ))

            if rota == "/relatorio":
                orgaos, empresas, statuses = buscar_filtros_relatorio()
                return self._responder_html(gerar_html_relatorio(orgaos, empresas, statuses, sessao))

            if rota in ("/relatorio/visualizar", "/relatorio/exportar-excel"):
                hoje_str = str(date.today())
                def _qs(k, default):
                    v = qs.get(k, [default])[0]
                    return v if v else default
                filtros = {
                    'data_inicio': _qs('data_inicio', str(date.today().replace(day=1))),
                    'data_fim':    _qs('data_fim',    hoje_str),
                    'orgao':       _qs('orgao',       ''),
                    'empresa':     _qs('empresa',     ''),
                    'status':      _qs('status',      ''),
                }
                for k in ('data_inicio', 'data_fim'):
                    try:
                        date.fromisoformat(filtros[k])
                    except ValueError:
                        filtros[k] = hoje_str
                por_orgao, processos = buscar_dados_relatorio(
                    filtros['data_inicio'], filtros['data_fim'],
                    filtros['orgao'] or None, filtros['empresa'] or None, filtros['status'] or None,
                )
                if rota == "/relatorio/visualizar":
                    return self._responder_html(
                        gerar_html_relatorio_visualizar(por_orgao, processos, filtros, sessao)
                    )
                dados_xlsx = _gerar_excel_relatorio(por_orgao, processos, filtros)
                nome = f"relatorio_{filtros['data_inicio']}_{filtros['data_fim']}.xlsx"
                return self._responder_arquivo(
                    dados_xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    nome,
                )

            return self._negar(404, "Página não encontrada.")

        except Exception as e:
            log.error(f"Erro ao processar GET {rota}: {e}")
            self._negar(500, "Erro interno. Consulte os logs do servidor.")

    # ── POST ───────────────────────────────────────────────────────
    def do_POST(self):
        parsed = urlparse(self.path)
        rota = parsed.path
        ip = self._ip_cliente()

        try:
            dados = self._ler_corpo_form()

            if rota == "/login":
                return self._processar_login(dados, ip)

            token, sessao = self._sessao_atual()
            if not sessao:
                return self._redirecionar("/login")

            # Rotas de API JSON não usam CSRF — a sessão via cookie já é suficiente
            if rota == "/api/admin/executar-monitoramento":
                if not sessao.get("is_admin"):
                    return self._negar()
                status = _ler_status_monitoramento()
                if status.get("running"):
                    return self._responder_json({"ok": False, "message": "Já está em execução"})
                subprocess.Popen([sys.executable, _SCRIPT_MANUAL], cwd=_BASE_DIR)
                return self._responder_json({"ok": True, "message": "Monitoramento iniciado"})

            if not auth.csrf_valido(token, dados.get("csrf", "")):
                return self._negar(400, "Sessão expirada, recarregue a página e tente novamente.")

            if rota == "/trocar-senha":
                return self._processar_trocar_senha(dados, sessao)

            if rota.startswith("/admin/usuarios"):
                if not sessao.get("is_admin"):
                    return self._negar()
                return self._processar_admin_usuarios(rota, dados, token, sessao)

            return self._negar(404, "Página não encontrada.")

        except Exception as e:
            log.error(f"Erro ao processar POST {rota}: {e}")
            self._negar(500, "Erro interno. Consulte os logs do servidor.")

    # ── Ações de POST ────────────────────────────────────────────────
    def _processar_login(self, dados, ip):
        email = (dados.get("email") or "").strip().lower()
        senha = dados.get("senha") or ""

        if auth.limite_tentativas_excedido(email, ip):
            return self._responder_html(
                gerar_html_login(erro="Muitas tentativas de login. Aguarde alguns minutos e tente novamente."),
                status=429,
            )

        usuario, erro = auth.autenticar_usuario(email, senha, ip)

        if erro:
            return self._responder_html(gerar_html_login(erro=erro), status=401)

        token = auth.gerar_token_sessao()
        auth_repository.criar_sessao(token, usuario["id"], ip)
        auth_repository.atualizar_ultimo_login(usuario["id"])

        destino = "/trocar-senha" if usuario.get("precisa_trocar_senha") else "/"
        self._redirecionar(destino, cookies=[auth.montar_cookie_sessao(token)])

    def _processar_trocar_senha(self, dados, sessao):
        usuario = auth_repository.buscar_usuario_por_id(sessao["usuario_id"])
        senha_atual = dados.get("senha_atual") or ""
        senha_nova = dados.get("senha_nova") or ""
        senha_confirma = dados.get("senha_confirma") or ""
        token_csrf = auth.gerar_token_csrf(auth.extrair_cookie(self.headers.get("Cookie", ""), auth.NOME_COOKIE_SESSAO))

        if not auth.verificar_senha(senha_atual, usuario["senha_hash"]):
            return self._responder_html(
                gerar_html_trocar_senha(token_csrf, erro="Senha atual incorreta.", obrigatorio=sessao.get("precisa_trocar_senha")),
                status=400,
            )

        if len(senha_nova) < 8 or senha_nova != senha_confirma:
            return self._responder_html(
                gerar_html_trocar_senha(token_csrf, erro="As senhas novas não conferem ou têm menos de 8 caracteres.", obrigatorio=sessao.get("precisa_trocar_senha")),
                status=400,
            )

        auth_repository.atualizar_senha(usuario["id"], auth.hash_senha(senha_nova), precisa_trocar_senha=False)
        self._redirecionar("/")

    def _processar_admin_usuarios(self, rota, dados, token, sessao):
        token_csrf = auth.gerar_token_csrf(token)

        if rota == "/admin/usuarios/criar":
            nome = (dados.get("nome") or "").strip()
            email = (dados.get("email") or "").strip().lower()
            senha_temp = dados.get("senha_temporaria") or ""
            is_admin = dados.get("is_admin") == "1"

            mensagem = f"Usuário {email} criado com sucesso."
            if not nome or not email or len(senha_temp) < 8:
                mensagem = "Preencha nome, e-mail e uma senha temporária com pelo menos 8 caracteres."
            else:
                try:
                    auth_repository.criar_usuario(nome, email, auth.hash_senha(senha_temp), is_admin=is_admin)
                except Exception:
                    mensagem = "Não foi possível criar o usuário — verifique se o e-mail já está cadastrado."

            usuarios = auth_repository.listar_usuarios()
            return self._responder_html(gerar_html_admin_usuarios(sessao, usuarios, token_csrf, mensagem=mensagem))

        partes = rota.rstrip("/").split("/")
        # ["", "admin", "usuarios", "<id>", "<acao>"]
        if len(partes) == 5 and partes[3].isdigit():
            usuario_id = int(partes[3])
            acao = partes[4]
            alvo = auth_repository.buscar_usuario_por_id(usuario_id)

            if not alvo:
                return self._negar(404, "Usuário não encontrado.")

            mensagem = None

            if acao == "status":
                auth_repository.atualizar_status(usuario_id, not alvo.get("ativo"))
                if alvo.get("ativo"):
                    auth_repository.invalidar_sessoes_do_usuario(usuario_id)
                mensagem = f"Usuário {alvo['email']} atualizado."

            elif acao == "resetar-senha":
                nova_senha = auth.gerar_token_sessao()[:12]
                auth_repository.atualizar_senha(usuario_id, auth.hash_senha(nova_senha), precisa_trocar_senha=True)
                auth_repository.invalidar_sessoes_do_usuario(usuario_id)
                mensagem = f"Nova senha temporária para {alvo['email']}: {nova_senha} (repasse com segurança — não fica salva em lugar nenhum)."

            elif acao == "excluir":
                if usuario_id == sessao.get("usuario_id"):
                    mensagem = "Você não pode excluir sua própria conta."
                else:
                    auth_repository.excluir_usuario(usuario_id)
                    mensagem = f"Usuário {alvo['email']} excluído permanentemente."

            usuarios = auth_repository.listar_usuarios()
            return self._responder_html(gerar_html_admin_usuarios(sessao, usuarios, token_csrf, mensagem=mensagem))

        return self._negar(404, "Ação inválida.")


def iniciar_dashboard():
    servidor = HTTPServer((HOST, PORTA), DashboardHandler)
    print("\n=== DASHBOARD SSA MONITOR PROCESSOS ===")
    print(f"Acesse: http://{HOST}:{PORTA}")
    print(f"Movimentações hoje: http://{HOST}:{PORTA}/movimentacoes-hoje")
    print("Pressione CTRL + C para encerrar.")
    servidor.serve_forever()


if __name__ == "__main__":
    iniciar_dashboard()
